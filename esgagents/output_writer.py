from __future__ import annotations

import json
import os
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Callable
from uuid import uuid4
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from esgagents.default_config import DEFAULT_CONFIG
from esgagents.publication import (
    PUBLICATION_STATUSES,
    apply_customer_answer_contract,
    customer_export_answer,
    resolved_publication_decision,
)
from esgagents.schemas import RunArtifacts, model_to_dict, validate_identifier
from esgagents.quality import QA_GRADES, resolved_answer_quality


AUDIT_COLUMNS = [
    "QID",
    "Source ID",
    "Category",
    "Question",
    "Answer Status",
    "RAG Pillar",
    "RAG Retrieval Confidence",
    "RAG Coverage",
    "RAG Answerable",
    "RAG Covered Facets",
    "RAG Missing Facets",
    "RAG Structured Coverage",
    "RAG Failure Code",
    "RAG Failure Reason",
    "RAG Retrieval Notes",
    "RAG Contract Violations",
    "RAG Contract Warnings",
    "Consumer Decision",
    "Upstream Hints",
    "Upstream Coverage Mismatch",
    "Local Evidence Accepted",
    "Local Acceptance Reason",
    "Metric Audit",
    "QA Grade",
    "Publication Status",
    "Publication Reason",
    "Publication Issues",
    "Final Answer",
    "Evidence Summary",
    "Sources",
    "QA Notes",
    "Agent Profile",
    "Quality Flags",
    "Revision Count",
    "Retrieval Attempts",
    "Skill Key",
    "Skill Name",
    "Skill Version",
    "Skill Source Path",
    "Skill Selection Reason",
    "Skill Checks",
    "Disclosure Flags",
    "Hard Failures",
    "Result Bucket",
    "Coverage Reason",
    "Coverage Issues",
    "Draft Answer",
    "Last Rejected Answer",
    "QA Failure Stage",
    "Sanitizer Actions",
]

COMBINED_QUALITATIVE_COLUMNS = [
    "EBX Indicator",
    "Status",
    "Field",
    "Original Evidence",
    "Final Answer",
]

QUANTITATIVE_COLUMNS = [
    "Metric ID",
    "Index",
    "Metric Name",
    "Value",
    "Unit",
    "Source",
    "Status",
    "Confidence",
    "Metadata",
]

RAG_METRIC_EVIDENCE_COLUMNS = [
    "QID",
    "Metric Status",
    "Metric Confidence",
    "Table Block",
    "Block Rank",
    "Block Role",
    "Entity",
    "Entity Class",
    "Metric Form",
    "Raw Evidence",
    "Parsed Facts",
    "Source Name",
    "Source Path",
    "Locator",
]

EXCEL_FALLBACK_TEXT = "The content contains characters that cannot be entered into Excel."
EXCEL_MAX_CELL_LENGTH = 32767
COMBINED_FILENAME_RE = re.compile(
    r"^\[langgraph\]\[.*\]report-(?P<date>\d{4}\.\d{2}\.\d{2})_(?P<number>\d+)\.xlsx$",
    re.IGNORECASE,
)
INVALID_FILENAME_CHARS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


class OutputPathError(ValueError):
    pass


class OutputRunExistsError(FileExistsError):
    pass


def _is_excel_text_char(char: str) -> bool:
    codepoint = ord(char)
    return (
        codepoint in (0x09, 0x0A, 0x0D)
        or 0x20 <= codepoint <= 0xD7FF
        or 0xE000 <= codepoint <= 0xFFFD
        or 0x10000 <= codepoint <= 0x10FFFF
    )


def clean_excel_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value

    cleaned = "".join(char for char in value if _is_excel_text_char(char))
    if cleaned.lstrip().startswith(("=", "+", "-", "@")):
        cleaned = "'" + cleaned
    return cleaned[:EXCEL_MAX_CELL_LENGTH]


def _metric_excel_value(value: Any) -> Any:
    raw = str(value or "").strip().replace(",", "").removesuffix("%")
    if not raw:
        return None
    try:
        number = Decimal(raw)
    except InvalidOperation:
        return value
    if abs(number) < Decimal("1e-12"):
        return 0
    if number.as_tuple().exponent < -6:
        number = number.quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP)
    if number == number.to_integral_value():
        return int(number)
    return float(number)


def _metric_period_sort_key(value: Any) -> tuple[tuple[int, Any], ...]:
    return tuple(
        (0, int(part)) if part.isdigit() else (1, part.casefold())
        for part in re.split(r"(\d+)", str(value or "").strip())
        if part
    )


def _metric_block_rank(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 10**9


def _is_total_metric(value: Any) -> bool:
    return bool(re.search(r"(?:합계|총계|소계|grand\s+total|sub\s*total|total)", str(value or ""), re.I))


class OutputWriter:
    def __init__(
        self,
        output_dir: str | Path,
        *,
        output_timezone: str = "Asia/Bangkok",
        now_provider: Callable[[], datetime] | None = None,
    ):
        self.output_dir = Path(output_dir)
        self.output_timezone = output_timezone
        self.now_provider = now_provider

    def load_existing(
        self,
        company_id: str,
        year: int,
        run_id: str,
    ) -> RunArtifacts | None:
        output_root = self.output_dir.resolve()
        safe_company_id = validate_identifier(str(company_id), "company_id")
        safe_run_id = validate_identifier(str(run_id), "run_id")
        company_root = output_root / safe_company_id
        candidates = [
            path
            for date_dir in company_root.glob("????_??_??")
            for path in [date_dir / safe_run_id]
        ]
        # Keep runs written before the date-directory layout readable for retries.
        candidates.append(company_root / str(int(year)) / safe_run_id)
        for run_dir in candidates:
            json_path = run_dir / "qualitative_run.json"
            if not json_path.is_file():
                continue
            try:
                with json_path.open("r", encoding="utf-8") as handle:
                    artifacts = RunArtifacts.model_validate(json.load(handle))
                if (
                    str(artifacts.company.get("company_id")) != safe_company_id
                    or int(artifacts.company.get("year")) != int(year)
                    or artifacts.run_id != safe_run_id
                ):
                    continue
                paths = [Path(value).resolve() for value in artifacts.output_paths.values()]
                if not paths or any(not path.is_file() for path in paths):
                    continue
                for path in paths:
                    path.relative_to(output_root)
                return artifacts
            except (OSError, ValueError, TypeError):
                continue
        return None

    def write(
        self,
        artifacts: RunArtifacts,
        *,
        retry_existing: bool = False,
    ) -> RunArtifacts:
        output_root = self.output_dir.resolve()
        company_id = validate_identifier(str(artifacts.company["company_id"]), "company_id")
        run_id = validate_identifier(artifacts.run_id, "run_id")
        _, run_dir = self._resolve_run_dir(company_id, int(artifacts.company["year"]), run_id)
        try:
            run_dir.mkdir(parents=True, exist_ok=False)
        except FileExistsError as exc:
            if not retry_existing:
                raise OutputRunExistsError("output run already exists") from exc
            existing = self.load_existing(
                company_id,
                int(artifacts.company["year"]),
                run_id,
            )
            if existing is not None:
                return existing
        for answer in artifacts.answers:
            apply_customer_answer_contract(answer)
        artifacts.stats = {bucket: 0 for bucket in ("answered", "empty", "weak", "failed")}
        for answer in artifacts.answers:
            bucket = _answer_result_bucket(answer)
            artifacts.stats[bucket if bucket in artifacts.stats else "empty"] += 1
        json_path = run_dir / "qualitative_run.json"
        coverage_path = run_dir / "coverage_summary.json"
        audit_json_path = run_dir / "qualitative_audit.json"
        existing_combined = next(
            (
                path
                for path in run_dir.glob("*.xlsx")
                if COMBINED_FILENAME_RE.match(path.name)
            ),
            None,
        )
        reservation_path: Path | None = None
        if existing_combined is not None:
            combined_path = existing_combined
        else:
            combined_path, reservation_path = self._reserve_combined_path(
                output_root=output_root,
                company_id=company_id,
                year=int(artifacts.company["year"]),
                company_name=str(artifacts.company.get("company_name") or company_id),
                run_dir=run_dir,
            )

        token = uuid4().hex
        json_temp = run_dir / f".qualitative_run.{token}.tmp"
        coverage_temp = run_dir / f".coverage_summary.{token}.tmp"
        audit_json_temp = run_dir / f".qualitative_audit.{token}.tmp"
        combined_temp = run_dir / f".combined_report.{token}.tmp.xlsx"
        try:
            artifacts.output_paths = {
                "json": str(json_path),
                "coverage_summary": str(coverage_path),
                "audit_json": str(audit_json_path),
                "combined_excel": str(combined_path),
            }
            with json_temp.open("w", encoding="utf-8") as handle:
                json.dump(model_to_dict(artifacts), handle, ensure_ascii=False, indent=2)
                handle.flush()
                os.fsync(handle.fileno())
            with coverage_temp.open("w", encoding="utf-8") as handle:
                json.dump(build_coverage_summary(artifacts), handle, ensure_ascii=False, indent=2)
                handle.flush()
                os.fsync(handle.fileno())
            with audit_json_temp.open("w", encoding="utf-8") as handle:
                json.dump(self._build_audit_json(artifacts), handle, ensure_ascii=False, indent=2)
                handle.flush()
                os.fsync(handle.fileno())
            self._write_combined_excel(artifacts, combined_temp)
            os.replace(audit_json_temp, audit_json_path)
            os.replace(combined_temp, combined_path)
            os.replace(coverage_temp, coverage_path)
            os.replace(json_temp, json_path)
            return artifacts
        finally:
            json_temp.unlink(missing_ok=True)
            coverage_temp.unlink(missing_ok=True)
            audit_json_temp.unlink(missing_ok=True)
            combined_temp.unlink(missing_ok=True)
            if reservation_path is not None:
                reservation_path.unlink(missing_ok=True)

    def _resolve_run_dir(
        self,
        company_id: str,
        year: int,
        run_id: str,
    ) -> tuple[Path, Path]:
        output_root = self.output_dir.resolve()
        safe_company_id = validate_identifier(str(company_id), "company_id")
        safe_run_id = validate_identifier(str(run_id), "run_id")
        run_dir = (output_root / safe_company_id / self._output_folder_date() / safe_run_id).resolve()
        legacy_run_dir = (output_root / safe_company_id / str(int(year)) / safe_run_id).resolve()
        if legacy_run_dir.is_dir() and not run_dir.exists():
            run_dir = legacy_run_dir
        try:
            comparable_root = os.path.normcase(os.path.normpath(str(output_root))).removeprefix("\\\\?\\")
            comparable_run = os.path.normcase(os.path.normpath(str(run_dir))).removeprefix("\\\\?\\")
            if os.path.commonpath([comparable_root, comparable_run]) != comparable_root:
                raise ValueError("path escaped output root")
        except ValueError as exc:
            raise OutputPathError("output run path must remain inside output_dir") from exc
        return output_root, run_dir

    def _build_audit_json(self, artifacts: RunArtifacts) -> dict[str, Any]:
        return {
            "columns": list(AUDIT_COLUMNS),
            "rows": [
                dict(zip(AUDIT_COLUMNS, self._audit_row(answer), strict=True))
                for answer in artifacts.answers
            ],
        }

    def _audit_row(self, answer: Any) -> list[Any]:
        quality = resolved_answer_quality(answer)
        return [
            answer.qid,
            answer.source_id,
            answer.category,
            answer.question,
            answer.answer_status,
            answer.rag_pillar,
            answer.rag_retrieval_confidence,
            answer.rag_coverage_status,
            answer.rag_answerable,
            "; ".join(answer.rag_covered_facets),
            "; ".join(answer.rag_missing_facets),
            json.dumps(answer.rag_coverage, ensure_ascii=False, sort_keys=True),
            answer.rag_failure_code,
            answer.rag_failure_reason,
            "; ".join(answer.rag_retrieval_notes),
            "; ".join(answer.rag_contract_violations),
            "; ".join(answer.rag_contract_warnings),
            answer.consumer_decision,
            json.dumps(answer.upstream_hints, ensure_ascii=False, sort_keys=True),
            answer.upstream_coverage_mismatch,
            answer.local_evidence_accepted,
            answer.local_acceptance_reason,
            json.dumps(answer.metric_audit, ensure_ascii=False, sort_keys=True),
            quality.grade,
            resolved_publication_decision(answer).status,
            resolved_publication_decision(answer).reason,
            "; ".join(resolved_publication_decision(answer).issues),
            _audit_display_answer(answer),
            answer.evidence_summary,
            "; ".join(_format_source(src) for src in answer.sources),
            "; ".join(answer.qa.notes),
            answer.agent_profile,
            "; ".join(answer.quality_flags),
            answer.revision_count,
            json.dumps(answer.retrieval_attempts, ensure_ascii=False, sort_keys=True),
            answer.skill_key,
            answer.skill_name,
            answer.skill_version,
            answer.skill_source_path,
            answer.skill_selection_reason,
            "; ".join(answer.skill_checks),
            "; ".join(answer.disclosure_flags),
            "; ".join(answer.hard_failures),
            _answer_result_bucket(answer),
            quality.reason,
            "; ".join(quality.issues),
            answer.draft_answer,
            answer.last_rejected_answer,
            answer.qa_failure_stage,
            "; ".join(answer.sanitizer_actions),
        ]

    def _write_excel(self, artifacts: RunArtifacts, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Qualitative Audit"
        self._append_excel_row(ws, AUDIT_COLUMNS)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")

        for answer in artifacts.answers:
            self._append_excel_row(ws, self._audit_row(answer))

        widths = [12, 16, 24, 44, 18, 18, 18, 18, 16, 36, 36, 60, 24, 52, 48, 52, 52, 22, 60, 22, 22, 30, 72, 14, 18, 28, 48, 60, 60, 52, 44, 20, 48, 16, 42, 16, 28, 16, 60, 32, 52, 44, 44, 16, 24, 24, 60, 60, 24, 48]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        self._write_rag_metric_evidence_sheet(wb, artifacts)
        wb.save(path)

    def _write_combined_excel(self, artifacts: RunArtifacts, path: Path) -> None:
        workbook = Workbook()
        qualitative = workbook.active
        qualitative.title = "Qualitative"

        self._append_excel_row(qualitative, COMBINED_QUALITATIVE_COLUMNS)
        for answer in artifacts.answers:
            self._append_excel_row(
                qualitative,
                [
                    answer.source_id,
                    _combined_status(answer),
                    _combined_field(answer),
                    answer.original_evidence,
                    customer_export_answer(answer),
                ],
            )

        self._style_report_sheet(
            qualitative,
            widths=[16, 22, 36, 72, 64],
            wrap_columns=set(range(1, 6)),
        )
        self._write_qualitative_table_metrics_sheet(workbook, artifacts)
        if artifacts.quantitative_results:
            quantitative = workbook.create_sheet("Quantitative")
            self._append_excel_row(quantitative, QUANTITATIVE_COLUMNS)
            for result in sorted(artifacts.quantitative_results, key=lambda item: item.index):
                self._append_excel_row(
                    quantitative,
                    [
                        result.metric_id,
                        result.index,
                        result.metric_name,
                        result.value,
                        result.unit,
                        result.source,
                        result.status,
                        result.confidence,
                        json.dumps(
                            result.metadata,
                            ensure_ascii=False,
                            sort_keys=True,
                            default=str,
                        ),
                    ],
                )
            self._style_report_sheet(
                quantitative,
                widths=[18, 10, 36, 18, 14, 38, 14, 14, 56],
                wrap_columns={3, 6, 9},
            )
        workbook.save(path)

    def _write_qualitative_table_metrics_sheet(
        self,
        workbook: Workbook,
        artifacts: RunArtifacts,
    ) -> None:
        worksheet = workbook.create_sheet("Qualitative Table Metrics")
        sections: list[dict[str, Any]] = []
        for answer in artifacts.answers:
            if answer.rag_metric_expected is not True:
                continue
            status = str(answer.rag_metric_status or "")
            audit = answer.metric_audit or {}
            absence = answer.rag_metric_absence or {}
            numeric_withheld = bool(audit.get("numeric_withheld"))
            facts = [
                fact
                for fact in (
                    audit.get("withheld_facts", [])
                    if numeric_withheld
                    else audit.get("accepted_facts", [])
                )
                if isinstance(fact, dict)
                and fact.get("block_role") == "primary"
                and bool(fact.get("entity_class") or fact.get("entity"))
            ]
            reason = str(absence.get("reason") or "")
            numeric_status = (
                "withheld_low_confidence"
                if numeric_withheld
                else "not_found"
                if status == "not_found"
                else "found_table_no_accepted_primary_fact"
            )
            if status != "found_table":
                facts = []

            grouped: dict[tuple[str, int, str, str], list[dict[str, Any]]] = {}
            for fact in facts:
                key = (
                    str(fact.get("table_block") or ""),
                    _metric_block_rank(fact.get("block_rank")),
                    str(fact.get("entity_class") or ""),
                    str(fact.get("entity") or ""),
                )
                grouped.setdefault(key, []).append(fact)

            if grouped:
                for key in sorted(grouped, key=lambda item: (item[1], item[2], item[3], item[0])):
                    table_block, block_rank, entity_class, entity = key
                    sections.append(
                        {
                            "answer": answer,
                            "table_block": table_block,
                            "block_rank": block_rank,
                            "entity_class": entity_class,
                            "entity": entity,
                            "facts": grouped[key],
                            "numeric_status": numeric_status if numeric_withheld else "accepted_primary",
                            "absence_reason": reason,
                            "withhold_values": numeric_withheld,
                        }
                    )
                continue

            sections.append(
                {
                    "answer": answer,
                    "table_block": "",
                    "block_rank": None,
                    "entity_class": "",
                    "entity": "",
                    "facts": [],
                    "numeric_status": numeric_status,
                    "absence_reason": reason,
                    "withhold_values": True,
                }
            )

        max_period_count = max(
            (
                len(
                    {
                        str(fact.get("period") or "").strip()
                        for fact in section["facts"]
                        if str(fact.get("period") or "").strip()
                    }
                )
                for section in sections
            ),
            default=1,
        )
        table_width = max(3, max_period_count + 2)
        section_counts: dict[str, int] = {}
        for section in sections:
            answer = section["answer"]
            section_counts[answer.qid] = section_counts.get(answer.qid, 0) + 1
            table_id = f"{answer.qid}-T{section_counts[answer.qid]:02d}"
            self._write_qualitative_metric_section(
                worksheet,
                section=section,
                table_id=table_id,
            )

        if not sections:
            worksheet["A1"] = "No qualitative table metrics are available."
            worksheet["A1"].font = Font(name="Carlito", size=11, italic=True, color="666666")

        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_properties.tabColor = "174A5A"
        worksheet.column_dimensions["A"].width = 46
        worksheet.column_dimensions["B"].width = 14
        for column in range(3, table_width + 1):
            worksheet.column_dimensions[get_column_letter(column)].width = 16
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    def _write_qualitative_metric_section(
        self,
        worksheet,
        *,
        section: dict[str, Any],
        table_id: str,
    ) -> None:
        answer = section["answer"]
        facts = section["facts"]
        periods = sorted(
            {
                str(fact.get("period") or "").strip()
                for fact in facts
                if str(fact.get("period") or "").strip()
            },
            key=_metric_period_sort_key,
        )
        section_width = max(3, len(periods) + 2)
        last_column = get_column_letter(section_width)
        title_row = worksheet.max_row + 1
        if title_row == 2 and worksheet["A1"].value is None:
            title_row = 1
        title = " | ".join(
            part
            for part in (
                table_id,
                answer.qid,
                answer.source_id,
                answer.question or answer.category,
            )
            if str(part or "").strip()
        )
        self._append_excel_row(worksheet, [title])
        worksheet.merge_cells(
            start_row=title_row,
            start_column=1,
            end_row=title_row,
            end_column=section_width,
        )
        title_range = worksheet[f"A{title_row}:{last_column}{title_row}"]
        for cell in title_range[0]:
            cell.fill = PatternFill("solid", fgColor="174A5A")
            cell.font = Font(name="Carlito", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        worksheet.row_dimensions[title_row].height = 27

        metadata_parts = []
        if section["table_block"]:
            metadata_parts.append(f"Table block: {section['table_block']}")
        if section["entity"]:
            metadata_parts.append(f"Entity: {section['entity']}")
        if section["entity_class"]:
            metadata_parts.append(f"Entity class: {section['entity_class']}")
        if section["block_rank"] is not None:
            metadata_parts.append(f"Block rank: {section['block_rank']}")
        metadata_parts.append(f"Numeric status: {section['numeric_status']}")
        if answer.rag_metric_confidence:
            metadata_parts.append(f"Metric confidence: {answer.rag_metric_confidence}")
        if section["absence_reason"]:
            metadata_parts.append(f"Absence reason: {section['absence_reason']}")
        metadata_row = worksheet.max_row + 1
        self._append_excel_row(
            worksheet,
            [" | ".join(metadata_parts)],
        )
        worksheet.merge_cells(
            start_row=metadata_row,
            start_column=1,
            end_row=metadata_row,
            end_column=section_width,
        )
        metadata_cell = worksheet.cell(metadata_row, 1)
        metadata_cell.font = Font(name="Carlito", size=10, bold=True, color="17313A")
        metadata_cell.fill = PatternFill("solid", fgColor="EAF1F4")
        metadata_cell.alignment = Alignment(vertical="center", wrap_text=True)
        worksheet.row_dimensions[metadata_row].height = 34

        header_row = worksheet.max_row + 1
        headers = ["Metric", "Unit", *(periods or ["Status"])]
        self._append_excel_row(worksheet, headers)
        header_range = worksheet[f"A{header_row}:{get_column_letter(len(headers))}{header_row}"]
        for cell in header_range[0]:
            cell.fill = PatternFill("solid", fgColor="174A5A")
            cell.font = Font(name="Carlito", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(vertical="center")
        worksheet.row_dimensions[header_row].height = 23

        metric_rows: dict[tuple[str, str], dict[str, Any]] = {}
        metric_order: dict[tuple[str, str], int] = {}
        for fact in facts:
            metric_key = (str(fact.get("metric") or ""), str(fact.get("unit") or ""))
            metric_order.setdefault(metric_key, len(metric_order))
            period = str(fact.get("period") or "").strip()
            if period:
                metric_rows.setdefault(metric_key, {})[period] = fact

        body_rows: list[list[Any]] = []
        for metric_key in sorted(
            metric_rows,
            key=lambda item: (0 if _is_total_metric(item[0]) else 1, metric_order[item]),
        ):
            metric, unit = metric_key
            period_facts = metric_rows[metric_key]
            values = []
            for period in periods:
                fact = period_facts.get(period)
                if fact is None or section["withhold_values"]:
                    values.append(None)
                    continue
                raw_value = fact.get("normalized_value")
                if raw_value in (None, ""):
                    raw_value = fact.get("value")
                values.append(_metric_excel_value(raw_value))
            body_rows.append([metric, unit, *values])

        if not body_rows:
            label = (
                "Numeric values withheld because metric confidence is low."
                if section["numeric_status"] == "withheld_low_confidence"
                else "No accepted primary metric was found."
            )
            body_rows = [[label, "", section["numeric_status"]]]

        body_start = worksheet.max_row + 1
        for values in body_rows:
            self._append_excel_row(worksheet, values)
        body_end = worksheet.max_row
        used_width = max(len(headers), max(len(row) for row in body_rows))
        border = Border(
            left=Side(style="thin", color="D2DDE2"),
            right=Side(style="thin", color="D2DDE2"),
            top=Side(style="thin", color="D2DDE2"),
            bottom=Side(style="thin", color="D2DDE2"),
        )
        for row in worksheet.iter_rows(
            min_row=body_start,
            max_row=body_end,
            min_col=1,
            max_col=used_width,
        ):
            for cell in row:
                cell.font = Font(name="Carlito", size=11)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="right" if cell.column >= 3 else "left",
                    vertical="center",
                    wrap_text=cell.column <= 2,
                )
                if cell.column >= 3 and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.######"
            worksheet.row_dimensions[row[0].row].height = 24

        worksheet.append([" "])
        spacer_row = worksheet.max_row
        worksheet.cell(spacer_row, 1).font = Font(color="FFFFFF")
        worksheet.row_dimensions[spacer_row].height = 12

    def _write_rag_metric_evidence_sheet(
        self,
        workbook: Workbook,
        artifacts: RunArtifacts,
    ) -> None:
        rows = [
            (answer, item)
            for answer in artifacts.answers
            for item in answer.rag_metric_evidence
            if isinstance(item, dict)
        ]
        if not rows:
            return
        worksheet = workbook.create_sheet("RAG Metric Evidence")
        self._append_excel_row(worksheet, RAG_METRIC_EVIDENCE_COLUMNS)
        for answer, item in rows:
            self._append_excel_row(
                worksheet,
                [
                    answer.qid,
                    answer.rag_metric_status,
                    answer.rag_metric_confidence,
                    item.get("table_block", ""),
                    item.get("block_rank"),
                    item.get("block_role", ""),
                    item.get("entity", ""),
                    item.get("entity_class", ""),
                    item.get("metric_form", ""),
                    item.get("raw_evidence_ko", ""),
                    json.dumps(item.get("facts", []), ensure_ascii=False, sort_keys=True),
                    item.get("source_name", ""),
                    item.get("source_path", ""),
                    json.dumps(item.get("locator", {}), ensure_ascii=False, sort_keys=True),
                ],
            )
        self._style_report_sheet(
            worksheet,
            widths=[12, 18, 18, 38, 12, 18, 22, 22, 16, 72, 60, 28, 42, 48],
            wrap_columns=set(range(1, 15)),
        )

    def _style_report_sheet(
        self,
        worksheet,
        *,
        widths: list[int],
        wrap_columns: set[int],
    ) -> None:
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 24
        for cell in worksheet[1]:
            cell.font = Font(name="Carlito", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="174A5A")
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.font = Font(name="Carlito", size=11)
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=cell.column in wrap_columns,
                )
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width

    def _reserve_combined_path(
        self,
        *,
        output_root: Path,
        company_id: str,
        year: int,
        company_name: str,
        run_dir: Path,
    ) -> tuple[Path, Path]:
        output_date = self._output_date()
        company_root = output_root / company_id
        reservation_dir = company_root / ".filename_reservations"
        reservation_dir.mkdir(parents=True, exist_ok=True)
        used_numbers = {
            number
            for path in company_root.rglob("*.xlsx")
            if (number := _combined_filename_number(path.name, output_date)) is not None
        }
        for lock_path in reservation_dir.glob(f"{output_date}_*.lock"):
            try:
                used_numbers.add(int(lock_path.stem.rsplit("_", 1)[-1]))
            except ValueError:
                continue
        number = max(used_numbers, default=0) + 1
        while True:
            reservation_path = reservation_dir / f"{output_date}_{number}.lock"
            try:
                descriptor = os.open(
                    reservation_path,
                    os.O_CREAT | os.O_EXCL | os.O_WRONLY,
                )
                os.close(descriptor)
                if any(
                    _combined_filename_number(path.name, output_date) == number
                    for path in company_root.rglob("*.xlsx")
                ):
                    reservation_path.unlink(missing_ok=True)
                    number += 1
                    continue
                break
            except FileExistsError:
                number += 1
        safe_company_name = sanitize_filename_component(company_name) or company_id
        filename = f"[langgraph][{safe_company_name}]report-{output_date}_{number}.xlsx"
        return run_dir / filename, reservation_path

    def _output_date(self) -> str:
        try:
            timezone = ZoneInfo(self.output_timezone)
        except Exception as exc:
            raise ValueError(f"invalid ESG_OUTPUT_TIMEZONE: {self.output_timezone}") from exc
        current = self.now_provider() if self.now_provider else datetime.now(timezone)
        if current.tzinfo is None:
            current = current.replace(tzinfo=timezone)
        else:
            current = current.astimezone(timezone)
        return current.strftime("%Y.%m.%d")

    def _output_folder_date(self) -> str:
        return self._output_date().replace(".", "_")

    def _append_excel_row(self, ws, values: list[Any]) -> None:
        is_empty_sheet = ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None
        row_idx = 1 if is_empty_sheet else ws.max_row + 1
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                cell.value = clean_excel_text(value)
            except Exception:
                cell.value = EXCEL_FALLBACK_TEXT


def _audit_display_answer(answer: Any) -> str:
    if str(getattr(answer, "final_answer", "") or "").strip():
        return str(answer.final_answer)
    if resolved_publication_decision(answer).status == "blocked":
        return str(getattr(answer, "last_rejected_answer", "") or "")
    return ""


def _answer_result_bucket(answer: Any) -> str:
    if getattr(answer, "result_bucket", None):
        return str(answer.result_bucket)
    if getattr(answer, "final_answer", ""):
        return "answered"
    qa = getattr(answer, "qa", None)
    if getattr(qa, "status", "") == "failed":
        return "failed"
    accepted_statuses = {str(s).lower() for s in DEFAULT_CONFIG["accepted_answer_statuses"]}
    answer_status = str(getattr(answer, "answer_status", "")).lower()
    qa_notes = " ".join(getattr(qa, "notes", []) or []).lower()
    if "weak" in qa_notes or (answer_status and answer_status not in accepted_statuses and answer_status != "missing"):
        return "weak"
    return "empty"


def build_coverage_summary(artifacts: RunArtifacts) -> dict[str, Any]:
    groups = {
        "empty_evidence": [],
        "metrics_missing": [],
        "draft_evidence": [],
        "draft_based_answers": [],
        "partial_answers": [],
        "missing_required_facets": [],
        "missing_expected_facets": [],
        "source_overstated": [],
        "unsupported_claim": [],
    }
    retrieval = {
        "retried_qids": [],
        "retry_helped_qids": [],
        "retry_unresolved_qids": [],
        "attempts_by_qid": {},
    }
    notes_counter: dict[str, int] = {}
    final_answer_stats = {"non_empty": 0, "empty": 0}
    qa_stats = {"passed": 0, "failed": 0, "empty": 0}
    quality_grade_stats = {grade: 0 for grade in QA_GRADES}
    quality_grade_qids = {grade: [] for grade in QA_GRADES}
    publication_status_stats = {status: 0 for status in PUBLICATION_STATUSES}
    publication_status_qids = {status: [] for status in PUBLICATION_STATUSES}
    publication_reason_stats: dict[str, int] = {}
    publication_reason_qids: dict[str, list[str]] = {}
    publication_issue_stats: dict[str, int] = {}
    publication_issue_qids: dict[str, list[str]] = {}
    coverage_reason_stats: dict[str, int] = {}
    coverage_reason_qids: dict[str, list[str]] = {}
    coverage_issue_stats: dict[str, int] = {}
    coverage_issue_qids: dict[str, list[str]] = {}
    coverage_matrix: dict[str, dict[str, int]] = {grade: {} for grade in QA_GRADES}
    rag_coverage_status_stats: dict[str, int] = {}
    rag_coverage_status_qids: dict[str, list[str]] = {}
    rag_answerable_stats = {"true": 0, "false": 0, "unknown": 0}
    rag_answerable_qids = {"true": [], "false": [], "unknown": []}
    rag_failure_code_stats: dict[str, int] = {}
    rag_failure_code_qids: dict[str, list[str]] = {}
    upstream_insufficient_qids: list[str] = []
    upstream_insufficient_failure_qids: dict[str, list[str]] = {}
    rag_contract_violation_qids: dict[str, list[str]] = {}
    rag_contract_warning_qids: dict[str, list[str]] = {}
    consumer_decision_stats: dict[str, int] = {}
    consumer_decision_qids: dict[str, list[str]] = {}
    consumer_funnel = {
        "total": len(artifacts.answers),
        "api_status_eligible": 0,
        "local_evidence_eligible": 0,
        "draft_non_empty": 0,
        "qa_passed": 0,
        "final_non_empty": 0,
    }
    metric_summary = {
        "qids_with_metric_rows": [],
        "metric_row_count": 0,
        "parsed_metric_row_count": 0,
        "malformed_metric_row_count": 0,
        "accepted_fact_count": 0,
        "conflict_count": 0,
        "conflict_qids": [],
        "all_numeric_facts_conflicted_qids": [],
        "status_qids": {},
        "absence_reason_qids": {},
        "low_confidence_qids": [],
        "summary_mismatch_qids": [],
        "summary_mismatch_fields": {},
        "contract_qids": {},
    }
    provenance_fallback_qids: list[str] = []
    upstream_mismatch_qids: list[str] = []
    salvaged_claim_qids: list[str] = []
    salvaged_claim_count = 0
    empty_final_answer_qids = []
    customer_answer_qids: list[str] = []
    review_exported_qids: list[str] = []
    rejected_candidate_qids: list[str] = []
    parity_mismatch_qids: list[str] = []
    local_admission_qids: dict[str, list[str]] = {}
    writer_fallback_qids: dict[str, list[str]] = {}
    for answer in artifacts.answers:
        qid = answer.qid
        decision = str(getattr(answer, "consumer_decision", "") or "blocked_evidence")
        consumer_decision_stats[decision] = consumer_decision_stats.get(decision, 0) + 1
        consumer_decision_qids.setdefault(decision, []).append(qid)
        if str(answer.answer_status or "").casefold() in {
            "high_confidence",
            "medium_confidence",
            "thin_but_usable",
        }:
            consumer_funnel["api_status_eligible"] += 1
        if getattr(answer, "local_evidence_accepted", False):
            consumer_funnel["local_evidence_eligible"] += 1
            reason = str(
                getattr(answer, "local_acceptance_reason", "") or "unknown"
            )
            local_admission_qids.setdefault(reason, []).append(qid)
        if answer.draft_answer:
            consumer_funnel["draft_non_empty"] += 1
        if str(getattr(answer.qa, "status", "") or "") == "passed":
            consumer_funnel["qa_passed"] += 1
        if answer.final_answer:
            final_answer_stats["non_empty"] += 1
            consumer_funnel["final_non_empty"] += 1
        else:
            final_answer_stats["empty"] += 1
            empty_final_answer_qids.append(qid)
        qa_status = str(getattr(answer.qa, "status", "") or "")
        if qa_status in qa_stats:
            qa_stats[qa_status] += 1
        quality = resolved_answer_quality(answer)
        quality_grade_stats[quality.grade] += 1
        quality_grade_qids[quality.grade].append(qid)
        publication = resolved_publication_decision(answer)
        publication_status_stats[publication.status] += 1
        publication_status_qids[publication.status].append(qid)
        exported_answer = customer_export_answer(answer)
        if exported_answer:
            customer_answer_qids.append(qid)
            if publication.status == "review_required":
                review_exported_qids.append(qid)
        if str(getattr(answer, "last_rejected_answer", "") or "").strip():
            rejected_candidate_qids.append(qid)
        if exported_answer != str(answer.final_answer or ""):
            parity_mismatch_qids.append(qid)
        publication_reason_stats[publication.reason] = (
            publication_reason_stats.get(publication.reason, 0) + 1
        )
        publication_reason_qids.setdefault(publication.reason, []).append(qid)
        for issue in publication.issues:
            publication_issue_stats[issue] = publication_issue_stats.get(issue, 0) + 1
            publication_issue_qids.setdefault(issue, []).append(qid)
        coverage_reason_stats[quality.reason] = coverage_reason_stats.get(quality.reason, 0) + 1
        coverage_reason_qids.setdefault(quality.reason, []).append(qid)
        grade_matrix = coverage_matrix[quality.grade]
        grade_matrix[quality.reason] = grade_matrix.get(quality.reason, 0) + 1
        for issue in quality.issues:
            coverage_issue_stats[issue] = coverage_issue_stats.get(issue, 0) + 1
            coverage_issue_qids.setdefault(issue, []).append(qid)
        rag_coverage = str(answer.rag_coverage_status or "unknown")
        rag_coverage_status_stats[rag_coverage] = rag_coverage_status_stats.get(rag_coverage, 0) + 1
        rag_coverage_status_qids.setdefault(rag_coverage, []).append(qid)
        answerable_key = (
            "true" if answer.rag_answerable is True
            else "false" if answer.rag_answerable is False
            else "unknown"
        )
        rag_answerable_stats[answerable_key] += 1
        rag_answerable_qids[answerable_key].append(qid)
        if answer.rag_failure_code:
            rag_failure_code_stats[answer.rag_failure_code] = rag_failure_code_stats.get(answer.rag_failure_code, 0) + 1
            rag_failure_code_qids.setdefault(answer.rag_failure_code, []).append(qid)
        if str(answer.answer_status or "").casefold() in {"insufficient", "no_evidence"}:
            upstream_insufficient_qids.append(qid)
            failure_code = str(answer.rag_failure_code or "UNKNOWN")
            upstream_insufficient_failure_qids.setdefault(failure_code, []).append(qid)
        for violation in answer.rag_contract_violations:
            rag_contract_violation_qids.setdefault(violation, []).append(qid)
        for warning in getattr(answer, "rag_contract_warnings", []) or []:
            rag_contract_warning_qids.setdefault(warning, []).append(qid)
        metric_audit = getattr(answer, "metric_audit", {}) or {}
        metric_status = str(getattr(answer, "rag_metric_status", "") or "legacy")
        metric_summary["status_qids"].setdefault(metric_status, []).append(qid)
        absence_reason = str(
            (getattr(answer, "rag_metric_absence", {}) or {}).get("reason") or ""
        )
        if absence_reason:
            metric_summary["absence_reason_qids"].setdefault(absence_reason, []).append(qid)
        if str(getattr(answer, "rag_metric_confidence", "") or "").casefold() == "low":
            metric_summary["low_confidence_qids"].append(qid)
        summary_mismatches = metric_audit.get("metric_summary_mismatches", {}) or {}
        if summary_mismatches:
            metric_summary["summary_mismatch_qids"].append(qid)
            for field in summary_mismatches:
                metric_summary["summary_mismatch_fields"].setdefault(field, []).append(qid)
        metric_contract = str(metric_audit.get("metric_contract") or "legacy")
        metric_summary["contract_qids"].setdefault(metric_contract, []).append(qid)
        metric_rows = int(metric_audit.get("metric_row_count", 0) or 0)
        if metric_rows:
            metric_summary["qids_with_metric_rows"].append(qid)
        for key in (
            "metric_row_count",
            "parsed_metric_row_count",
            "malformed_metric_row_count",
            "accepted_fact_count",
            "conflict_count",
        ):
            metric_summary[key] += int(metric_audit.get(key, 0) or 0)
        if int(metric_audit.get("conflict_count", 0) or 0):
            metric_summary["conflict_qids"].append(qid)
        if metric_audit.get("all_numeric_facts_conflicted"):
            metric_summary["all_numeric_facts_conflicted_qids"].append(qid)
        if any(source.get("provenance_fallback") for source in answer.sources):
            provenance_fallback_qids.append(qid)
        if getattr(answer, "upstream_coverage_mismatch", False):
            upstream_mismatch_qids.append(qid)
        salvage_actions = [
            action for action in (answer.sanitizer_actions or [])
            if str(action).startswith("removed_claim:")
        ]
        if salvage_actions:
            salvaged_claim_qids.append(qid)
            salvaged_claim_count += len(salvage_actions)
        notes = list(getattr(answer.qa, "notes", []) or [])
        flags = list(answer.quality_flags or [])
        for fallback_flag in (
            "non_substantive_llm_output",
            "structured_metric_fallback",
            "unsupported_metric_llm_output",
        ):
            if fallback_flag in flags:
                writer_fallback_qids.setdefault(fallback_flag, []).append(qid)
        checks = list(answer.skill_checks or [])
        combined = " | ".join([*notes, *flags, *checks]).casefold()
        for note in notes:
            notes_counter[note] = notes_counter.get(note, 0) + 1
        if "empty evidence" in combined:
            groups["empty_evidence"].append(qid)
        if (
            "metric_result" in combined
            or "reporting_period" in combined
            or "missing_quantitative_metric_result" in combined
        ):
            groups["metrics_missing"].append(qid)
        if any(term in combined for term in ("draft", "proposal", "planned", "under review", "검토", "계획")):
            groups["draft_evidence"].append(qid)
        if answer.final_answer and "draft_based_answer" in flags:
            groups["draft_based_answers"].append(qid)
        if answer.final_answer and ("partial_answer" in flags or "missing data disclosed" in combined):
            groups["partial_answers"].append(qid)
        if "missing required facet:" in combined:
            groups["missing_required_facets"].append(qid)
        if "missing facet:" in combined:
            groups["missing_expected_facets"].append(qid)
        if "source usage overstated" in combined:
            groups["source_overstated"].append(qid)
        if "unsupported numeric claim" in combined or "unsupported certification" in combined:
            groups["unsupported_claim"].append(qid)
        attempts = list(answer.retrieval_attempts or [])
        if attempts:
            retrieval["attempts_by_qid"][qid] = attempts
        retry_attempts = [
            attempt for attempt in attempts
            if str(attempt.get("retry_reason", "")) not in {"", "initial"}
        ]
        if retry_attempts:
            retrieval["retried_qids"].append(qid)
            retry_improved = _retry_improved(attempts)
            if retry_improved is True or (retry_improved is None and answer.final_answer):
                retrieval["retry_helped_qids"].append(qid)
            else:
                retrieval["retry_unresolved_qids"].append(qid)

    return {
        "run_id": artifacts.run_id,
        "company": artifacts.company,
        "total_qids": len(artifacts.answers),
        "stats": dict(artifacts.stats),
        "bucket_stats": dict(artifacts.stats),
        "final_answer_stats": final_answer_stats,
        "qa_stats": qa_stats,
        "quality_grade_stats": quality_grade_stats,
        "quality_grade_qids": {
            grade: sorted(qids) for grade, qids in quality_grade_qids.items()
        },
        "publication_status_stats": publication_status_stats,
        "publication_status_qids": {
            status: sorted(qids) for status, qids in publication_status_qids.items()
        },
        "publication_reason_stats": dict(sorted(publication_reason_stats.items())),
        "publication_reason_qids": {
            reason: sorted(qids)
            for reason, qids in sorted(publication_reason_qids.items())
        },
        "publication_issue_stats": dict(sorted(publication_issue_stats.items())),
        "publication_issue_qids": {
            issue: sorted(qids)
            for issue, qids in sorted(publication_issue_qids.items())
        },
        "customer_answer_count": len(customer_answer_qids),
        "customer_answer_qids": sorted(customer_answer_qids),
        "review_exported_count": len(review_exported_qids),
        "review_exported_qids": sorted(review_exported_qids),
        "rejected_candidate_count": len(rejected_candidate_qids),
        "rejected_candidate_qids": sorted(rejected_candidate_qids),
        "json_xlsx_answer_parity": not parity_mismatch_qids,
        "json_xlsx_answer_parity_mismatch_qids": sorted(parity_mismatch_qids),
        "coverage_reason_stats": dict(sorted(coverage_reason_stats.items())),
        "coverage_reason_qids": {
            reason: sorted(qids) for reason, qids in sorted(coverage_reason_qids.items())
        },
        "coverage_issue_stats": dict(sorted(coverage_issue_stats.items())),
        "coverage_issue_qids": {
            issue: sorted(qids) for issue, qids in sorted(coverage_issue_qids.items())
        },
        "coverage_matrix": {
            grade: dict(sorted(reasons.items())) for grade, reasons in coverage_matrix.items()
        },
        "provenance": dict(artifacts.provenance),
        "rag": {
            "coverage_status_stats": dict(sorted(rag_coverage_status_stats.items())),
            "coverage_status_qids": {
                status: sorted(qids) for status, qids in sorted(rag_coverage_status_qids.items())
            },
            "answerable_stats": rag_answerable_stats,
            "answerable_qids": {
                status: sorted(qids) for status, qids in rag_answerable_qids.items()
            },
            "failure_code_stats": dict(sorted(rag_failure_code_stats.items())),
            "failure_code_qids": {
                code: sorted(qids) for code, qids in sorted(rag_failure_code_qids.items())
            },
            "upstream_insufficient": {
                "count": len(set(upstream_insufficient_qids)),
                "qids": sorted(set(upstream_insufficient_qids)),
                "failure_code_qids": {
                    code: sorted(set(qids))
                    for code, qids in sorted(upstream_insufficient_failure_qids.items())
                },
            },
            "contract_violation_count": sum(
                len(qids) for qids in rag_contract_violation_qids.values()
            ) + sum(len(trace.contract_violations) for trace in artifacts.rag_request_traces),
            "contract_violation_qids": {
                violation: sorted(qids)
                for violation, qids in sorted(rag_contract_violation_qids.items())
            },
            "contract_warning_count": sum(
                len(qids) for qids in rag_contract_warning_qids.values()
            ),
            "contract_warning_qids": {
                warning: sorted(qids)
                for warning, qids in sorted(rag_contract_warning_qids.items())
            },
            "request_ids": sorted({trace.request_id for trace in artifacts.rag_request_traces if trace.request_id}),
            "api_versions": sorted({trace.api_version for trace in artifacts.rag_request_traces if trace.api_version}),
            "rag_versions": sorted({trace.rag_version for trace in artifacts.rag_request_traces if trace.rag_version}),
            "index_versions": sorted({trace.index_version for trace in artifacts.rag_request_traces if trace.index_version}),
            "request_traces": [model_to_dict(trace) for trace in artifacts.rag_request_traces],
        },
        "consumer_funnel": consumer_funnel,
        "consumer_decision_stats": dict(sorted(consumer_decision_stats.items())),
        "consumer_decision_qids": {
            decision: sorted(qids)
            for decision, qids in sorted(consumer_decision_qids.items())
        },
        "local_admission": {
            "count": len(
                {
                    qid
                    for qids in local_admission_qids.values()
                    for qid in qids
                }
            ),
            "reason_qids": {
                reason: sorted(set(qids))
                for reason, qids in sorted(local_admission_qids.items())
            },
        },
        "writer_fallback": {
            "count": len(
                {
                    qid
                    for qids in writer_fallback_qids.values()
                    for qid in qids
                }
            ),
            "reason_qids": {
                reason: sorted(set(qids))
                for reason, qids in sorted(writer_fallback_qids.items())
            },
        },
        "metric_facts": {
            **metric_summary,
            "qids_with_metric_rows": sorted(set(metric_summary["qids_with_metric_rows"])),
            "conflict_qids": sorted(set(metric_summary["conflict_qids"])),
            "all_numeric_facts_conflicted_qids": sorted(
                set(metric_summary["all_numeric_facts_conflicted_qids"])
            ),
            "status_qids": {
                status: sorted(set(qids))
                for status, qids in sorted(metric_summary["status_qids"].items())
            },
            "absence_reason_qids": {
                reason: sorted(set(qids))
                for reason, qids in sorted(metric_summary["absence_reason_qids"].items())
            },
            "low_confidence_qids": sorted(set(metric_summary["low_confidence_qids"])),
            "summary_mismatch_count": len(
                set(metric_summary["summary_mismatch_qids"])
            ),
            "summary_mismatch_qids": sorted(
                set(metric_summary["summary_mismatch_qids"])
            ),
            "summary_mismatch_fields": {
                field: sorted(set(qids))
                for field, qids in sorted(
                    metric_summary["summary_mismatch_fields"].items()
                )
            },
            "contract_qids": {
                contract: sorted(set(qids))
                for contract, qids in sorted(metric_summary["contract_qids"].items())
            },
        },
        "provenance_fallback": {
            "count": len(set(provenance_fallback_qids)),
            "qids": sorted(set(provenance_fallback_qids)),
        },
        "upstream_mismatch": {
            "count": len(set(upstream_mismatch_qids)),
            "qids": sorted(set(upstream_mismatch_qids)),
        },
        "claim_salvage": {
            "claim_count": salvaged_claim_count,
            "qid_count": len(set(salvaged_claim_qids)),
            "qids": sorted(set(salvaged_claim_qids)),
        },
        "empty_final_answer_qids": sorted(empty_final_answer_qids),
        "top_failure_notes": [
            {"note": note, "count": count}
            for note, count in sorted(notes_counter.items(), key=lambda item: (-item[1], item[0]))[:20]
        ],
        "groups": {name: sorted(set(qids)) for name, qids in groups.items()},
        "retrieval": {
            "retried_qids": sorted(set(retrieval["retried_qids"])),
            "retry_helped_qids": sorted(set(retrieval["retry_helped_qids"])),
            "retry_unresolved_qids": sorted(set(retrieval["retry_unresolved_qids"])),
            "attempts_by_qid": retrieval["attempts_by_qid"],
        },
    }


def coverage_reason(answer: Any) -> str:
    return resolved_answer_quality(answer).reason


def sanitize_filename_component(value: str) -> str:
    cleaned = INVALID_FILENAME_CHARS_RE.sub("_", str(value or ""))
    cleaned = re.sub(r"\s+", " ", cleaned).strip().rstrip(". ")
    return cleaned[:160].rstrip(". ")


def _combined_filename_number(filename: str, output_date: str) -> int | None:
    match = COMBINED_FILENAME_RE.match(filename)
    if not match or match.group("date") != output_date:
        return None
    return int(match.group("number"))


def _retry_improved(attempts: list[dict[str, Any]]) -> bool | None:
    result_attempts = [
        attempt
        for attempt in attempts
        if str(attempt.get("answer_status", "") or "")
    ]
    if len(result_attempts) < 2:
        return None
    status_rank = {
        "high_confidence": 5,
        "medium_confidence": 4,
        "thin_but_usable": 3,
        "insufficient": 2,
        "no_evidence": 1,
    }
    metric_rank = {"found_table": 2, "not_found": 1, "not_expected": 0}

    def quality(attempt: dict[str, Any]) -> tuple[int, int, int, int, int, float]:
        return (
            status_rank.get(str(attempt.get("answer_status", "") or "").casefold(), 0),
            metric_rank.get(str(attempt.get("metric_status", "") or "").casefold(), 0),
            int(attempt.get("metric_primary_block_count", 0) or 0),
            len(attempt.get("covered_facets", []) or []),
            int(attempt.get("eligible_item_count", 0) or 0),
            float(attempt.get("retrieval_confidence", 0.0) or 0.0),
        )

    return quality(result_attempts[-1]) > quality(result_attempts[0])


def _combined_status(answer: Any) -> str:
    publication = resolved_publication_decision(answer)
    answer_status = {
        "published": "PUBLISHED",
        "review_required": "REVIEW",
        "blocked": "BLOCKED",
    }.get(str(publication.status or "").casefold(), str(publication.status or "UNKNOWN").upper())
    return (
        f"Answer: {answer_status}\n"
        f"Evidence: {_combined_evidence_status(answer)}"
    )


def _combined_evidence_status(answer: Any) -> str:
    flags = {str(flag).casefold() for flag in getattr(answer, "quality_flags", []) or []}
    coverage_issues = {
        str(issue).casefold() for issue in getattr(answer, "coverage_issues", []) or []
    }
    notes = {
        str(note).casefold()
        for note in getattr(getattr(answer, "qa", None), "notes", []) or []
    }
    combined = flags | coverage_issues | notes
    metric_audit = getattr(answer, "metric_audit", {}) or {}
    answer_status = str(getattr(answer, "answer_status", "") or "").casefold()
    rag_coverage = str(getattr(answer, "rag_coverage_status", "") or "").casefold()
    publication = resolved_publication_decision(answer)

    if (
        getattr(answer, "upstream_coverage_mismatch", False)
        or "upstream_coverage_mismatch" in combined
        or "metric_summary_mismatch" in combined
        or "conflicting_metric" in combined
        or bool(metric_audit.get("metric_summary_mismatches"))
        or int(metric_audit.get("conflict_count", 0) or 0) > 0
    ):
        return "MISMATCH"
    if (
        str(getattr(answer, "rag_metric_confidence", "") or "").casefold() == "low"
        or "metric_low_confidence" in combined
        or "metric_numeric_withheld" in combined
        or bool(metric_audit.get("numeric_withheld"))
    ):
        return "METRIC_LOW_CONFIDENCE"
    if (
        getattr(answer, "rag_metric_expected", None) is True
        or str(getattr(answer, "rag_metric_status", "") or "").casefold()
        in {"found_table", "not_found"}
        or "metric_not_found" in combined
        or "malformed_metric_row" in combined
        or "missing_metric_or_period" in combined
    ):
        if "metric_not_found" in combined or str(
            getattr(answer, "rag_metric_status", "") or ""
        ).casefold() == "not_found":
            return "METRIC_REVIEW"
        if publication.status != "published":
            return "METRIC_REVIEW"
    if (
        answer_status in {"", "missing", "insufficient", "no_evidence"}
        or rag_coverage in {"insufficient", "no_evidence"}
        or bool(getattr(answer, "rag_failure_code", "") or "")
        or bool(getattr(answer, "rag_contract_violations", []) or [])
    ):
        return "ERROR"
    if (
        answer_status == "thin_but_usable"
        or rag_coverage == "partial"
        or "partial_answer" in combined
        or "rag_partial_coverage" in combined
        or "local_partial_evidence" in combined
        or "thin_evidence" in combined
        or publication.status == "review_required"
    ):
        return "PARTIAL"
    return "SUFFICIENT"


def _combined_field(answer: Any) -> str:
    return " / ".join(
        part
        for part in (
            getattr(answer, "area", "") or getattr(answer, "category", ""),
            _display_pillar(getattr(answer, "rag_pillar", "")),
            getattr(answer, "question", ""),
        )
        if part
    )


def _display_pillar(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "(" in text and ")" in text:
        return text
    return {
        "strategy": "전략 (Strategy)",
        "governance": "거버넌스 (Governance)",
        "risk_management": "위험 관리 (Risk Management)",
        "risk management": "위험 관리 (Risk Management)",
        "metrics": "지표 (Metrics)",
    }.get(text.casefold(), text)


def _format_source(source: dict[str, Any]) -> str:
    locator = source.get("locator") if isinstance(source.get("locator"), dict) else {}
    metadata = "; ".join(
        f"{field}={source.get(field, '')}"
        for field in (
            "document_id",
            "chunk_id",
            "canonical_source_id",
            "provenance_key",
            "provenance_method",
            "provenance_fallback",
            "source_tier",
            "source_type",
            "document_status",
            "document_version",
            "effective_date",
            "topic",
            "subtopic",
            "semantic_label",
            "semantic_score",
            "reranker_score",
            "vector_score",
            "score",
            "classification_reason",
        )
        if source.get(field) is not None and source.get(field) != ""
    )
    locator_text = ",".join(
        f"{field}={value}"
        for field, value in locator.items()
        if value is not None and value != ""
    )
    if locator_text:
        metadata = "; ".join(filter(None, [metadata, f"locator({locator_text})"]))
    location = " | ".join(
        str(value)
        for value in (
            source.get("source_name"),
            source.get("source_path"),
            source.get("provenance_key"),
        )
        if value
    )
    return f"[{metadata}] {location}" if metadata else location
