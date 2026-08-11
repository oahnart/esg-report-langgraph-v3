import json
from pathlib import Path

from openpyxl import load_workbook

from esgagents.graph.esg_graph import ESGQualitativeGraph
from esgagents.rag_client import TeamRagClient


def test_graph_runs_subset_with_mock_rag_without_writing_outputs(tmp_path):
    def transport(endpoint, payload, timeout):
        return {
            "company_id": payload["company_id"],
            "results": [
                {
                    "question_id": qid,
                    "question_ko": f"{qid} question",
                    "normalized_answer_ko": f"{qid} normalized answer",
                    "answer_status": "high_confidence",
                    "items": [
                        {
                            "score": 1,
                            "raw_evidence_ko": f"{qid} normalized answer",
                            "source_name": "source.docx",
                            "source_path": "ESG/source.docx",
                            "semantic_label": "strong",
                            "semantic_score": 0.95,
                        }
                    ],
                }
                for qid in payload.get("question_ids", payload.get("item_ids", []))
            ],
        }

    graph = ESGQualitativeGraph(
        config={
            "output_dir": str(tmp_path),
            "team_rag_batch_size": 2,
            "agent_mode": "offline",
            "quantitative_output_enabled": False,
            "quantitative_input_mode": "file",
            "quantitative_input_dir": str(tmp_path / "inputs"),
        },
        rag_client=TeamRagClient(
            "mock://rag",
            transport=transport,
            qualitative_path="/qualitative/evidence/v2",
        ),
    )
    artifacts = graph.generate(
        {
            "company_id": "iljinhysolus",
            "company_name": "Iljin Hysolus",
            "year": 2025,
            "scale": "large",
            "industry": "TC",
            "item_ids": ["Q001", "Q016", "Q047"],
        },
        write_outputs=False,
    )

    assert artifacts.stats == {"answered": 2, "empty": 0, "weak": 0, "failed": 1}
    assert [answer.qid for answer in artifacts.answers] == ["Q001", "Q016", "Q047"]
    assert artifacts.answers[0].sources[0]["source_name"] == "source.docx"
    assert artifacts.answers[0].skill_name
    assert artifacts.answers[2].qa.status == "failed"
    assert "missing required facet: metric_result" in artifacts.answers[2].qa.notes
    assert artifacts.quantitative_stats == {}
    assert artifacts.quantitative_results == []


def test_graph_runs_with_checkpoint_enabled(tmp_path):
    def transport(endpoint, payload, timeout):
        return {
            "company_id": payload["company_id"],
            "results": [
                {
                    "question_id": qid,
                    "question_ko": qid,
                    "normalized_answer_ko": "answer",
                    "answer_status": "high_confidence",
                    "items": [
                        {
                            "raw_evidence_ko": "answer",
                            "source_name": "doc",
                            "source_path": "ESG/doc.docx",
                            "semantic_label": "strong",
                            "semantic_score": 0.9,
                        }
                    ],
                }
                for qid in payload.get("question_ids", payload.get("item_ids", []))
            ],
        }

    graph = ESGQualitativeGraph(
        config={
            "checkpoint_enabled": True,
            "cache_dir": str(tmp_path),
            "output_dir": str(tmp_path),
            "agent_mode": "offline",
            "quantitative_input_mode": "file",
            "quantitative_input_dir": str(tmp_path / "inputs"),
        },
        rag_client=TeamRagClient(
            "mock://rag",
            transport=transport,
            qualitative_path="/qualitative/evidence/v2",
        ),
    )
    artifacts = graph.generate(
        {
            "company_id": "checkpoint_company",
            "company_name": "Checkpoint Company",
            "year": 2025,
            "scale": "large",
            "industry": "TC",
            "item_ids": ["Q001"],
            "run_id": "run_checkpoint",
        },
        write_outputs=False,
    )

    assert artifacts.stats["answered"] == 1
    assert (tmp_path / "checkpoints" / "CHECKPOINT_COMPANY.db").exists()


def test_graph_uses_v3_metric_items_without_quantitative_loader(tmp_path):
    def transport(endpoint, payload, timeout):
        return _v3_metric_response(payload)

    class Loader:
        def load(self, company):
            raise AssertionError("quantitative loader should not be called")

    graph = ESGQualitativeGraph(
        config={
            "output_dir": str(tmp_path),
            "agent_mode": "offline",
            "semantic_qa_enabled": False,
            "quantitative_output_enabled": False,
            "quantitative_input_mode": "api",
            "quantitative_api_base_url": "https://quant.example",
        },
        rag_client=TeamRagClient("mock://rag", transport=transport),
    )
    graph.agents.quantitative_agent.input_loader = Loader()

    artifacts = graph.generate(
        {
            "company_id": "daewoong",
            "company_name": "Daewoong",
            "year": 2025,
            "scale": "large",
            "industry": "HC",
            "item_ids": ["Q031"],
            "run_id": "run_v3_metric",
        },
    )

    answer = artifacts.answers[0]
    assert artifacts.quantitative_results == []
    assert artifacts.quantitative_stats == {}
    assert "Scope 1 emissions" in answer.final_answer
    assert "2025" in answer.final_answer
    assert "100.0 tCO2e" in answer.final_answer
    assert answer.sources[0]["semantic_label"] == "metric_row"
    assert answer.sources[0]["locator"]["spans_units"] == ["tCO2e"]
    workbook = load_workbook(Path(artifacts.output_paths["combined_excel"]), read_only=True)
    assert workbook.sheetnames == ["Qualitative", "Qualitative Table Metrics"]


def test_quantitative_output_does_not_overwrite_v3_qualitative_metric_evidence(tmp_path):
    def transport(endpoint, payload, timeout):
        return _v3_metric_response(payload)

    class Loader:
        def load(self, company):
            return (
                {
                    "items": [
                        {
                            "metric_id": "QUANT-0001",
                            "mapped_qualitative_qid": "Q031",
                            "metric_name": "Scope 1 emissions",
                            "value": 999,
                            "unit": "tCO2e",
                            "source": "quantitative.xlsx",
                        }
                    ]
                },
                "quantitative_raw.json",
            )

    graph = ESGQualitativeGraph(
        config={
            "output_dir": str(tmp_path / "outputs"),
            "agent_mode": "offline",
            "semantic_qa_enabled": False,
            "quantitative_output_enabled": True,
        },
        rag_client=TeamRagClient("mock://rag", transport=transport),
    )
    graph.agents.quantitative_agent.input_loader = Loader()

    artifacts = graph.generate(
        {
            "company_id": "daewoong",
            "company_name": "Daewoong",
            "year": 2025,
            "scale": "large",
            "industry": "HC",
            "item_ids": ["Q031"],
            "run_id": "run_v3_metric_with_quant",
        }
    )

    answer = artifacts.answers[0]
    assert artifacts.quantitative_results[0].value == 999
    assert "Scope 1 emissions" in answer.final_answer
    assert "2025" in answer.final_answer
    assert "100.0 tCO2e" in answer.final_answer
    assert "quantitative_metric_bridge" not in answer.quality_flags
    assert "999" not in answer.evidence_summary
    workbook = load_workbook(Path(artifacts.output_paths["combined_excel"]), read_only=True)
    assert workbook.sheetnames == [
        "Qualitative",
        "Qualitative Table Metrics",
        "Quantitative",
    ]


def test_graph_writes_combined_workbook_with_quantitative_input(tmp_path):
    def transport(endpoint, payload, timeout):
        return {
            "company_id": payload["company_id"],
            "results": [
                {
                    "question_id": qid,
                    "question_ko": qid,
                    "normalized_answer_ko": "answer",
                    "answer_status": "high_confidence",
                    "items": [
                        {
                            "raw_evidence_ko": "evidence",
                            "source_name": "report.pdf",
                            "source_path": "ESG/report.pdf",
                            "semantic_label": "strong",
                            "semantic_score": 0.9,
                        }
                    ],
                }
                for qid in payload.get("question_ids", payload.get("item_ids", []))
            ],
        }

    input_root = tmp_path / "inputs"
    quantitative_path = input_root / "combined_company" / "2025" / "quantitative_raw.json"
    quantitative_path.parent.mkdir(parents=True)
    quantitative_path.write_text(
        json.dumps(
            {
                "metrics": [
                    {
                        "metric_id": "QUANT-0001",
                        "value": 100,
                        "unit": "people",
                        "source": "report.pdf",
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    graph = ESGQualitativeGraph(
        config={
            "output_dir": str(tmp_path / "outputs"),
            "quantitative_input_mode": "file",
            "quantitative_input_dir": str(input_root),
            "quantitative_output_enabled": True,
            "agent_mode": "offline",
            "output_timezone": "Asia/Bangkok",
        },
        rag_client=TeamRagClient(
            "mock://rag",
            transport=transport,
            qualitative_path="/qualitative/evidence/v2",
        ),
    )

    artifacts = graph.generate(
        {
            "company_id": "combined_company",
            "company_name": "Combined Company",
            "year": 2025,
            "scale": "large",
            "industry": "TC",
            "item_ids": ["Q001"],
            "run_id": "run_combined",
        }
    )

    assert artifacts.quantitative_stats == {"total": 251, "filled": 1, "missing": 250}
    assert artifacts.quantitative_results[0].value == 100
    combined_path = Path(artifacts.output_paths["combined_excel"])
    assert combined_path.exists()
    workbook = load_workbook(combined_path, read_only=True)
    assert workbook.sheetnames == [
        "Qualitative",
        "Qualitative Table Metrics",
        "Quantitative",
    ]
    assert workbook["Quantitative"].max_row == 252


def test_graph_writes_quant_210_workbook_and_withholds_confirmation_values(tmp_path):
    def transport(endpoint, payload, timeout):
        return {
            "company_id": payload["company_id"],
            "results": [
                {
                    "question_id": qid,
                    "question_ko": qid,
                    "normalized_answer_ko": "answer",
                    "answer_status": "high_confidence",
                    "items": [
                        {
                            "raw_evidence_ko": "evidence",
                            "source_name": "report.pdf",
                            "source_path": "ESG/report.pdf",
                            "semantic_label": "strong",
                            "semantic_score": 0.9,
                        }
                    ],
                }
                for qid in payload.get("question_ids", payload.get("item_ids", []))
            ],
        }

    item_ids = [str(value) for value in range(596, 806)]
    item_ids[-1] = "814"
    answered = 0
    items = []
    for item_id in item_ids:
        if item_id in {"693", "814"}:
            status = "needs_confirmation"
            value = 0.7209302325581395 if item_id == "693" else 5.071467
            reason = "needs confirmation"
        elif answered < 76:
            status = "answered"
            answered += 1
            value = answered
            reason = "guided_cell"
        else:
            status = "missing"
            value = None
            reason = "missing"
        items.append(
            {
                "item_id": item_id,
                "mapped_qualitative_qid": "Q063" if item_id == "596" else "",
                "source_id": "EBX-Q-063" if item_id == "596" else "",
                "domain": "사회",
                "category": "구성원 및 다양성",
                "subcategory": "성별",
                "item": f"metric {item_id}",
                "question": f"question {item_id}",
                "unit": "명",
                "standards": {},
                "answer": {
                    "value": value,
                    "unit": "명",
                    "year": 2025,
                    "status": status,
                    "confidence": "high",
                    "reason": reason,
                    "source": "정량데이터.xlsx",
                    "evidence": [{"source": "정량데이터.xlsx", "cell": "J1"}] if value else [],
                },
            }
        )

    class Loader:
        def load(self, company):
            return (
                {
                    "company_id": company.company_id,
                    "company_name": company.company_name,
                    "year": company.year,
                    "kind": "quantitative",
                    "catalog_pack": "quant_210",
                    "total": 210,
                    "items": items,
                },
                "api_snapshot_quantitative.json",
            )

    graph = ESGQualitativeGraph(
        config={
            "output_dir": str(tmp_path / "outputs"),
            "agent_mode": "offline",
            "output_timezone": "Asia/Bangkok",
            "quantitative_output_enabled": True,
        },
        rag_client=TeamRagClient(
            "mock://rag",
            transport=transport,
            qualitative_path="/qualitative/evidence/v2",
        ),
    )
    graph.agents.quantitative_agent.input_loader = Loader()

    artifacts = graph.generate(
        {
            "company_id": "daewoong",
            "company_name": "대웅제약",
            "year": 2025,
            "scale": "large",
            "industry": "HC",
            "item_ids": ["Q063"],
            "run_id": "run_quant_210",
        }
    )

    assert artifacts.quantitative_stats == {
        "total": 210,
        "filled": 76,
        "missing": 132,
        "needs_confirmation": 2,
        "published": 76,
    }
    assert len(artifacts.quantitative_results) == 210
    assert artifacts.answers[0].qid == "Q063"
    assert "quantitative_metric_bridge" not in artifacts.answers[0].quality_flags
    assert artifacts.answers[0].evidence_summary == "evidence"
    combined_path = Path(artifacts.output_paths["combined_excel"])
    workbook = load_workbook(combined_path, read_only=True)
    quantitative = workbook["Quantitative"]
    assert quantitative.max_row == 211
    rows_by_metric = {
        row[0]: row
        for row in quantitative.iter_rows(min_row=2, values_only=True)
        if row[0] in {"693", "814"}
    }
    assert rows_by_metric["693"][3] is None
    assert rows_by_metric["814"][3] is None
    assert rows_by_metric["693"][6] == "missing"
    assert "needs_confirmation" in rows_by_metric["693"][8]


def _v3_metric_response(payload):
    return {
        "company_id": payload["company_id"],
        "request_id": "rag-v3-metric",
        "api_version": "3.0",
        "rag_version": "rag-v3",
        "index_version": "daewoong_20260805",
        "generated_at": "2026-08-07T09:04:39+07:00",
        "latency_ms": 10,
        "warnings": [],
        "results": [
            {
                "question_id": qid,
                "question_ko": qid,
                "pillar": "metrics",
                "normalized_answer_ko": "Scope 1 emissions | tCO2e | 2025=100.0",
                "answer_status": "medium_confidence",
                "retrieval_confidence": 0.9,
                "coverage_status": "complete",
                "answerable": True,
                "covered_facets": ["metric_result", "reporting_period"],
                "missing_facets": [],
                "failure_code": None,
                "failure_reason": "",
                "retrieval_notes": ["metric evidence mirrored in items"],
                "coverage": {
                    "direct_answer": True,
                    "supports_metric_result": True,
                    "supports_reporting_period": True,
                },
                "items": [
                    {
                        "score": 1,
                        "vector_score": None,
                        "reranker_score": None,
                        "semantic_score": 1,
                        "semantic_label": "metric_row",
                        "semantic_reason": "metric_lane_row",
                        "raw_evidence_ko": "Scope 1 emissions | tCO2e | 2025=100.0",
                        "source_name": "metrics.xlsx",
                        "source_path": "ESG/metrics.xlsx",
                        "document_id": "metric_lane::metrics.xlsx",
                        "chunk_id": "scope1::2025",
                        "canonical_source_id": "metric_lane::metrics.xlsx",
                        "source_type": "operational_record",
                        "document_status": "approved",
                        "source_tier": "tier_2_operational",
                        "document_version": None,
                        "effective_date": None,
                        "topic": "metrics",
                        "subtopic": "metric_lane",
                        "locator": {
                            "sheet_name": "GHG",
                            "section": "Scope 1 emissions",
                            "cell_range": "A1:B1",
                            "spans_units": ["tCO2e"],
                            "confidence": "exact",
                        },
                    }
                ],
            }
            for qid in payload.get("question_ids", payload.get("item_ids", []))
        ],
    }
