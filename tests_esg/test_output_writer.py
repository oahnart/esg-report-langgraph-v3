import json
import re
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from esgagents.output_writer import (
    AUDIT_COLUMNS,
    COMBINED_QUALITATIVE_COLUMNS,
    EXCEL_FALLBACK_TEXT,
    OutputRunExistsError,
    OutputWriter,
    QUANTITATIVE_COLUMNS,
    _combined_evidence_status,
    _metric_excel_value,
    build_coverage_summary,
    clean_excel_text,
)
from esgagents.agents.managers.report_manager import ReportManagerAgent
from esgagents.schemas import (
    AnswerRecord,
    ClaimSupport,
    EvidenceItem,
    MetricEvidenceItem,
    QAResult,
    QuantitativeResult,
    RagRequestTrace,
    RunArtifacts,
)


def _audit_value(worksheet, column_name, row=2):
    column = AUDIT_COLUMNS.index(column_name) + 1
    return worksheet.cell(row=row, column=column).value


def test_writer_original_evidence_uses_raw_accepted_writer_items():
    metric_item = MetricEvidenceItem(
        raw_evidence_ko="Metric row | tCO2e | 2025=10",
        block_role="primary",
        table_block="GHG",
        entity_class="company",
    )
    narrative_item = EvidenceItem(
        raw_evidence_ko="  Original narrative evidence.\nKeep this raw spacing.  "
    )

    result = ReportManagerAgent._writer_original_evidence(
        {
            "metric_items": [metric_item],
            "narrative_items": [narrative_item],
            "metric_audit": {"metric_status": "found_table"},
        }
    )

    assert result == narrative_item.raw_evidence_ko


def _audit_json_value(payload, column_name, row=0):
    return payload["rows"][row][column_name]


@pytest.mark.parametrize(
    ("answer", "expected"),
    [
        (
            AnswerRecord(
                qid="Q001",
                answer_status="high_confidence",
                rag_coverage_status="complete",
                final_answer="answer",
                qa=QAResult(status="passed"),
                qa_grade="full",
                publication_status="published",
            ),
            "SUFFICIENT",
        ),
        (
            AnswerRecord(
                qid="Q002",
                answer_status="thin_but_usable",
                rag_coverage_status="partial",
                final_answer="answer",
                qa=QAResult(status="passed"),
                qa_grade="partial",
                quality_flags=["partial_answer"],
            ),
            "PARTIAL",
        ),
        (
            AnswerRecord(
                qid="Q003",
                answer_status="insufficient",
                rag_coverage_status="insufficient",
                qa=QAResult(status="failed"),
            ),
            "ERROR",
        ),
        (
            AnswerRecord(
                qid="Q003A",
                answer_status="insufficient",
                rag_coverage_status="insufficient",
                final_answer="There is still a safe direction for review.",
                qa=QAResult(status="passed"),
            ),
            "PARTIAL",
        ),
        (
            AnswerRecord(
                qid="Q004",
                answer_status="high_confidence",
                upstream_coverage_mismatch=True,
                qa=QAResult(status="passed"),
            ),
            "MISMATCH",
        ),
        (
            AnswerRecord(
                qid="Q005",
                answer_status="high_confidence",
                rag_metric_confidence="low",
                qa=QAResult(status="passed"),
                quality_flags=["metric_low_confidence"],
            ),
            "METRIC_LOW_CONFIDENCE",
        ),
        (
            AnswerRecord(
                qid="Q006",
                answer_status="high_confidence",
                rag_metric_expected=True,
                rag_metric_status="not_found",
                qa=QAResult(status="passed"),
                quality_flags=["metric_not_found"],
            ),
            "METRIC_REVIEW",
        ),
    ],
)
def test_combined_evidence_status_uses_review_categories(answer, expected):
    assert _combined_evidence_status(answer) == expected


def test_output_writer_persists_resolved_quality_contract(tmp_path):
    artifacts = RunArtifacts(
        run_id="run_quality_contract",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 1},
        stats={"answered": 1, "empty": 0, "weak": 0, "failed": 0},
        answers=[
            AnswerRecord(
                qid="Q019",
                source_id="EBX-Q-019",
                category="정보보호",
                question="개인정보 침해 및 정보보안 사고 현황",
                answer_status="high_confidence",
                rag_coverage_status="complete",
                rag_answerable=True,
                final_answer="대웅제약은 정보보안 체계를 지속 강화하고 있습니다.",
                evidence_summary="evidence",
                sources=[
                    {
                        "source_name": "metric.xlsx",
                        "source_path": "ESG/metric.xlsx",
                        "source_tier": "tier_2_operational",
                    }
                ],
                qa=QAResult(status="passed", notes=["semantic review passed"]),
                quality_flags=["metric_low_confidence"],
                qa_grade="full",
                coverage_reason="complete_grounded_answer",
                coverage_issues=[],
            )
        ],
    )

    written = OutputWriter(tmp_path).write(artifacts)
    payload = json.loads(Path(written.output_paths["json"]).read_text(encoding="utf-8"))
    audit = json.loads(Path(written.output_paths["audit_json"]).read_text(encoding="utf-8"))
    summary = json.loads(Path(written.output_paths["coverage_summary"]).read_text(encoding="utf-8"))

    answer = payload["answers"][0]
    assert answer["qa_grade"] == "partial"
    assert answer["coverage_reason"] == "qa_invariant_violation"
    assert "qa_invariant_violation" in answer["coverage_issues"]
    assert answer["publication_status"] == "review_required"
    assert _audit_json_value(audit, "QA Grade") == "partial"
    assert summary["quality_grade_stats"]["partial"] == 1
    assert summary["quality_grade_qids"]["partial"] == ["Q019"]


def test_metric_excel_value_rounds_display_precision_without_decimal_artifacts():
    assert _metric_excel_value("0.000000000000000000") == 0
    assert _metric_excel_value("12771.822287491") == 12771.822287
    assert _metric_excel_value("10.500000") == 10.5


def test_output_writer_creates_json_and_excel_audit(tmp_path):
    artifacts = RunArtifacts(
        run_id="run_test",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 1},
        stats={"answered": 1, "empty": 0, "weak": 0, "failed": 0},
        answers=[
            AnswerRecord(
                qid="Q001",
                source_id="EBX-Q-001",
                area="일반",
                category="ESG",
                question="question",
                answer_status="high_confidence",
                rag_pillar="governance",
                rag_retrieval_confidence=0.94,
                rag_coverage_status="complete",
                rag_answerable=True,
                rag_covered_facets=["accountable_body", "role"],
                rag_coverage={"direct_answer": True, "supports_role": True},
                draft_answer="draft answer",
                final_answer="answer",
                last_rejected_answer="rejected answer",
                qa_failure_stage="skill_policy_critic",
                sanitizer_actions=["removed_unsupported_numeric_claim:30%"],
                evidence_summary="evidence",
                sources=[{
                    "source_name": "doc.docx",
                    "source_path": "ESG/doc.docx",
                    "document_id": "doc_123",
                    "chunk_id": "chunk_123",
                    "canonical_source_id": "src_123",
                    "source_tier": "tier_1_governing",
                    "source_type": "policy_procedure",
                    "document_status": "governing",
                    "effective_date": "2025-01-01",
                    "classification_reason": "inferred_keyword:policy",
                    "locator": {"page": 3, "section": "Governance"},
                }],
                qa=QAResult(status="passed", notes=["grounded"]),
                claim_support=[
                    ClaimSupport(
                        claim_id="c1",
                        claim_text="answer",
                        source_ids=["src_123"],
                        support_tier="tier_1_governing",
                        support_status="grounded",
                        facets=["accountable_body", "role"],
                        reporting_period="2025",
                    )
                ],
                retrieval_attempts=[
                    {"top_k": 5, "retry_reason": "initial", "eligible_item_count": 0},
                    {"top_k": 12, "retry_reason": "empty evidence", "eligible_item_count": 1},
                ],
            )
        ],
        rag_request_traces=[
            RagRequestTrace(
                request_id="rag-req-1",
                api_version="3.0",
                rag_version="rag-v3",
                index_version="index-v1",
                generated_at=datetime(2025, 8, 1, tzinfo=ZoneInfo("UTC")),
                requested_item_ids=["Q001"],
                top_k=5,
            )
        ],
    )

    written = OutputWriter(tmp_path).write(artifacts)
    audit = json.loads(Path(written.output_paths["audit_json"]).read_text(encoding="utf-8"))

    assert audit["columns"] == AUDIT_COLUMNS
    assert len(audit["rows"]) == 1
    assert audit["rows"][0]["QID"] == "Q001"
    assert _audit_json_value(audit, "QA Grade") == "full"
    assert _audit_json_value(audit, "Result Bucket") == "answered"
    assert _audit_json_value(audit, "Coverage Reason") == "complete_grounded_answer"
    assert "empty evidence" in _audit_json_value(audit, "Retrieval Attempts")
    assert _audit_json_value(audit, "Draft Answer") == "draft answer"
    assert _audit_json_value(audit, "Last Rejected Answer") == "rejected answer"
    assert _audit_json_value(audit, "QA Failure Stage") == "skill_policy_critic"
    assert _audit_json_value(audit, "Sanitizer Actions") == "removed_unsupported_numeric_claim:30%"
    assert "source_tier=tier_1_governing" in _audit_json_value(audit, "Sources")
    assert "classification_reason=inferred_keyword:policy" in _audit_json_value(audit, "Sources")
    assert _audit_json_value(audit, "RAG Coverage") == "complete"
    assert _audit_json_value(audit, "RAG Answerable") is True
    assert '"direct_answer": true' in _audit_json_value(audit, "RAG Structured Coverage")
    assert "locator(page=3,section=Governance)" in _audit_json_value(audit, "Sources")
    combined = load_workbook(written.output_paths["combined_excel"])["Qualitative"]
    assert combined["E2"].value == "answer"
    payload = json.loads(open(written.output_paths["json"], encoding="utf-8").read())
    assert payload["answers"][0]["draft_answer"] == "draft answer"
    assert payload["answers"][0]["last_rejected_answer"] == "rejected answer"
    assert payload["answers"][0]["qa_failure_stage"] == "skill_policy_critic"
    assert payload["answers"][0]["sanitizer_actions"] == ["removed_unsupported_numeric_claim:30%"]
    assert payload["answers"][0]["sources"][0]["source_tier"] == "tier_1_governing"
    assert payload["answers"][0]["claim_support"] == [
        {
            "claim_id": "c1",
            "claim_text": "answer",
            "source_ids": ["src_123"],
            "support_tier": "tier_1_governing",
            "support_status": "grounded",
            "facets": ["accountable_body", "role"],
            "reporting_period": "2025",
            "attribution_required": False,
        }
    ]
    assert payload["rag_request_traces"][0]["generated_at"] == "2025-08-01T00:00:00Z"
    assert written.output_paths["json"].endswith("qualitative_run.json")
    assert written.output_paths["coverage_summary"].endswith("coverage_summary.json")
    coverage = json.loads(open(written.output_paths["coverage_summary"], encoding="utf-8").read())
    assert coverage["total_qids"] == 1
    assert coverage["groups"]["unsupported_claim"] == []
    assert coverage["retrieval"]["retried_qids"] == ["Q001"]
    assert coverage["retrieval"]["retry_helped_qids"] == ["Q001"]
    assert coverage["rag"]["coverage_status_stats"] == {"complete": 1}
    assert coverage["rag"]["answerable_stats"]["true"] == 1
    assert coverage["rag"]["request_ids"] == ["rag-req-1"]


def test_output_writer_adds_full_rag_metric_evidence_sheet(tmp_path):
    artifacts = RunArtifacts(
        run_id="run_metric_sheet",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 1},
        stats={"answered": 1, "empty": 0, "weak": 0, "failed": 0},
        answers=[
            AnswerRecord(
                qid="Q039",
                answer_status="medium_confidence",
                rag_metric_expected=True,
                rag_metric_status="found_table",
                metric_audit={
                    "metric_status": "found_table",
                    "accepted_facts": [
                        {
                            "table_block": "Water > Pharm",
                            "block_rank": 1,
                            "block_role": "primary",
                            "entity": "Daewoong Pharm",
                            "entity_class": "daewoong_pharm",
                            "metric": "Water use",
                            "period": "2025",
                            "value": "10",
                            "normalized_value": "10",
                            "unit": "ton",
                            "value_role": "actual",
                            "source_name": "metrics.xlsx",
                            "source_id": "ESG/metrics.xlsx",
                        }
                    ],
                },
                rag_metric_evidence=[
                    {
                        "table_block": "Water > Pharm",
                        "block_rank": 1,
                        "block_role": "primary",
                        "entity": "Daewoong Pharm",
                        "entity_class": "daewoong_pharm",
                        "metric_form": "table_row",
                        "raw_evidence_ko": "Water use | ton | 2025=10",
                        "facts": [{"metric": "Water use", "period": "2025", "value": "10"}],
                        "source_name": "metrics.xlsx",
                        "source_path": "ESG/metrics.xlsx",
                        "locator": {"sheet_name": "Water", "cell_range": "A1:D1"},
                    },
                    {
                        "table_block": "Water > Pharm",
                        "block_rank": 1,
                        "block_role": "denominator",
                        "entity": "Daewoong Pharm",
                        "entity_class": "daewoong_pharm",
                        "metric_form": "table_row",
                        "raw_evidence_ko": "Sales | KRW | 2025=100",
                        "facts": [],
                    },
                ],
                final_answer="Water use was reported.",
                qa=QAResult(status="passed"),
            )
        ],
    )

    written = OutputWriter(tmp_path).write(artifacts)
    combined_book = load_workbook(written.output_paths["combined_excel"])

    assert combined_book.sheetnames == ["Qualitative", "Qualitative Table Metrics"]
    audit = json.loads(Path(written.output_paths["audit_json"]).read_text(encoding="utf-8"))
    assert "Water use" in _audit_json_value(audit, "Metric Audit")
    customer_metric_sheet = combined_book["Qualitative Table Metrics"]
    assert customer_metric_sheet["A1"].value.startswith("Q039-T01 | Q039")
    assert "Table block: Water > Pharm" in customer_metric_sheet["A2"].value
    assert "Entity: Daewoong Pharm" in customer_metric_sheet["A2"].value
    assert "Numeric status: accepted_primary" in customer_metric_sheet["A2"].value
    assert [customer_metric_sheet.cell(3, column).value for column in range(1, 4)] == [
        "Metric",
        "Unit",
        "2025",
    ]
    assert [customer_metric_sheet.cell(4, column).value for column in range(1, 4)] == [
        "Water use",
        "ton",
        10,
    ]


def test_coverage_summary_tracks_metric_summary_mismatch_fields():
    artifacts = RunArtifacts(
        run_id="run_metric_mismatch",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 1},
        stats={"answered": 1, "empty": 0, "weak": 0, "failed": 0},
        answers=[
            AnswerRecord(
                qid="Q039",
                answer_status="medium_confidence",
                rag_metric_status="found_table",
                metric_audit={
                    "metric_contract": "new",
                    "metric_status": "found_table",
                    "metric_summary_mismatches": {
                        "n_rows": {"expected": 23, "actual": 22},
                        "n_blocks": {"expected": 5, "actual": 4},
                    },
                },
                final_answer="Water use was reported.",
                quality_flags=["metric_summary_mismatch", "human_review_required"],
                qa=QAResult(status="passed"),
            )
        ],
    )

    coverage = build_coverage_summary(artifacts)

    assert coverage["metric_facts"]["summary_mismatch_count"] == 1
    assert coverage["metric_facts"]["summary_mismatch_qids"] == ["Q039"]
    assert coverage["metric_facts"]["summary_mismatch_fields"] == {
        "n_blocks": ["Q039"],
        "n_rows": ["Q039"],
    }


def test_customer_metric_sheet_respects_status_roles_and_low_confidence(tmp_path):
    primary_fact = {
        "table_block": "Water > Pharm",
        "block_rank": 1,
        "block_role": "primary",
        "entity": "Daewoong Pharm",
        "entity_class": "daewoong_pharm",
        "metric": "Water use",
        "period": "2025",
        "value": "10.5",
        "normalized_value": "10.5",
        "unit": "ton",
        "value_role": "actual",
    }
    scope_fact = {
        **primary_fact,
        "block_role": "scope_variant",
        "entity": "Factory A",
        "entity_class": "factory",
        "value": "3",
        "normalized_value": "3",
    }
    artifacts = RunArtifacts(
        run_id="run_customer_metric_contract",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "v1", "question_count": 4},
        stats={"answered": 4, "empty": 0, "weak": 0, "failed": 0},
        answers=[
            AnswerRecord(
                qid="Q001",
                source_id="EBX-Q-001",
                rag_metric_expected=False,
                rag_metric_status="not_expected",
                final_answer="Legacy qualitative answer.",
                qa=QAResult(status="passed"),
            ),
            AnswerRecord(
                qid="Q039",
                source_id="EBX-Q-039",
                rag_metric_expected=True,
                rag_metric_status="found_table",
                metric_audit={
                    "metric_status": "found_table",
                    "accepted_facts": [primary_fact, scope_fact],
                },
                final_answer="Narrative water-management explanation.",
                qa=QAResult(status="passed"),
            ),
            AnswerRecord(
                qid="Q019",
                source_id="EBX-Q-019",
                rag_metric_expected=True,
                rag_metric_status="found_table",
                rag_metric_confidence="low",
                metric_audit={
                    "metric_status": "found_table",
                    "numeric_withheld": True,
                    "accepted_facts": [primary_fact],
                    "withheld_facts": [primary_fact],
                },
                final_answer="Narrative incident-response explanation.",
                qa=QAResult(status="passed"),
            ),
            AnswerRecord(
                qid="Q095",
                source_id="EBX-Q-095",
                rag_metric_expected=True,
                rag_metric_status="not_found",
                rag_metric_absence={"reason": "below_threshold"},
                metric_audit={"metric_status": "not_found", "accepted_facts": []},
                final_answer="Narrative stakeholder-engagement explanation.",
                qa=QAResult(status="passed"),
            ),
        ],
    )

    written = OutputWriter(tmp_path).write(artifacts)
    sheet = load_workbook(written.output_paths["combined_excel"])[
        "Qualitative Table Metrics"
    ]
    title_rows = {
        str(sheet.cell(row, 1).value).split("-T", 1)[0]: row
        for row in range(1, sheet.max_row + 1)
        if re.match(r"^Q\d{3}-T\d{2}\s+\|", str(sheet.cell(row, 1).value or ""))
    }

    assert "Q001" not in title_rows

    q039 = title_rows["Q039"]
    assert "Entity: Daewoong Pharm" in sheet.cell(q039 + 1, 1).value
    assert "Numeric status: accepted_primary" in sheet.cell(q039 + 1, 1).value
    assert sheet.cell(q039 + 2, 3).value == "2025"
    assert sheet.cell(q039 + 3, 1).value == "Water use"
    assert sheet.cell(q039 + 3, 3).value == 10.5

    q019 = title_rows["Q019"]
    assert "Numeric status: withheld_low_confidence" in sheet.cell(q019 + 1, 1).value
    assert sheet.cell(q019 + 3, 1).value == "Water use"
    assert sheet.cell(q019 + 3, 3).value is None

    q095 = title_rows["Q095"]
    assert "Absence reason: below_threshold" in sheet.cell(q095 + 1, 1).value
    assert sheet.cell(q095 + 2, 3).value == "Status"
    assert sheet.cell(q095 + 3, 3).value == "not_found"


def test_customer_metric_sheet_renders_each_block_and_entity_as_an_independent_table(tmp_path):
    facts = [
        {
            "table_block": "Water > Pharm",
            "block_rank": 1,
            "block_role": "primary",
            "entity": "Daewoong Pharm",
            "entity_class": "daewoong_pharm",
            "metric": "Water use",
            "period": "2024",
            "normalized_value": "9",
            "unit": "ton",
        },
        {
            "table_block": "Water > Pharm",
            "block_rank": 1,
            "block_role": "primary",
            "entity": "Daewoong Pharm",
            "entity_class": "daewoong_pharm",
            "metric": "Total",
            "period": "2025",
            "normalized_value": "12",
            "unit": "ton",
        },
        {
            "table_block": "Water > Group",
            "block_rank": 2,
            "block_role": "primary",
            "entity": "Daewoong Group",
            "entity_class": "group_total",
            "metric": "Water use",
            "period": "2025",
            "normalized_value": "20",
            "unit": "ton",
        },
    ]
    artifacts = RunArtifacts(
        run_id="run_independent_metric_tables",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "v1", "question_count": 1},
        stats={"answered": 1, "empty": 0, "weak": 0, "failed": 0},
        answers=[
            AnswerRecord(
                qid="Q039",
                source_id="EBX-Q-039",
                question="Water metrics",
                rag_metric_expected=True,
                rag_metric_status="found_table",
                metric_audit={"metric_status": "found_table", "accepted_facts": facts},
                final_answer="Narrative water-management explanation.",
                qa=QAResult(status="passed"),
            )
        ],
    )

    written = OutputWriter(tmp_path).write(artifacts)
    sheet = load_workbook(written.output_paths["combined_excel"])[
        "Qualitative Table Metrics"
    ]
    title_rows = [
        row
        for row in range(1, sheet.max_row + 1)
        if re.match(r"^Q039-T\d{2}\s+\|", str(sheet.cell(row, 1).value or ""))
    ]

    assert len(title_rows) == 2
    first, second = title_rows
    assert sheet.cell(first, 1).value.startswith("Q039-T01 | Q039 | EBX-Q-039")
    assert "Entity: Daewoong Pharm" in sheet.cell(first + 1, 1).value
    assert [sheet.cell(first + 2, column).value for column in range(1, 5)] == [
        "Metric",
        "Unit",
        "2024",
        "2025",
    ]
    assert sheet.cell(first + 3, 1).value == "Total"
    assert sheet.cell(first + 3, 4).value == 12
    assert "Entity: Daewoong Group" in sheet.cell(second + 1, 1).value
    assert sheet.cell(second + 3, 1).value == "Water use"
    assert sheet.cell(second + 3, 3).value == 20


def test_output_writer_cleans_illegal_excel_characters(tmp_path):
    artifacts = RunArtifacts(
        run_id="run_test",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 1},
        stats={"answered": 1, "empty": 0, "weak": 0, "failed": 0},
        answers=[
            AnswerRecord(
                qid="Q025",
                source_id="EBX-Q-025",
                category="ESG",
                question="question",
                answer_status="high_confidence",
                final_answer="answer",
                evidence_summary="감사실\x01소개",
                sources=[{"source_name": "doc\x01.docx", "source_path": "ESG/doc.docx"}],
                qa=QAResult(status="passed", notes=["grounded\x01note"]),
            )
        ],
    )

    written = OutputWriter(tmp_path).write(artifacts)
    audit = json.loads(Path(written.output_paths["audit_json"]).read_text(encoding="utf-8"))

    assert _audit_json_value(audit, "Evidence Summary") == artifacts.answers[0].evidence_summary
    assert "ESG/doc.docx" in _audit_json_value(audit, "Sources")
    assert _audit_json_value(audit, "QA Notes") == artifacts.answers[0].qa.notes[0]


def test_clean_excel_text_removes_invalid_xml_characters_and_truncates():
    assert clean_excel_text("ok\ud800bad\uffff") == "okbad"
    assert len(clean_excel_text("x" * 40000)) == 32767


@pytest.mark.parametrize("value", ["=1+1", "+CMD", "-2+3", "@SUM(A1:A2)", "  =1+1"])
def test_clean_excel_text_neutralizes_formula_prefixes(value):
    assert clean_excel_text(value) == "'" + value


def test_excel_formula_neutralization_does_not_change_json(tmp_path):
    artifacts = RunArtifacts(
        run_id="run_formula",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 1},
        stats={"answered": 1, "empty": 0, "weak": 0, "failed": 0},
        answers=[
                AnswerRecord(
                    qid="Q001",
                    answer_status="high_confidence",
                    final_answer="=1+1",
                    evidence_summary="@SUM(A1:A2)",
                    sources=[
                        {
                            "source_name": "source.pdf",
                            "source_path": "ESG/source.pdf",
                            "canonical_source_id": "src_formula",
                            "source_tier": "tier_1_governing",
                            "document_status": "governing",
                        }
                    ],
                    claim_support=[
                        ClaimSupport(
                            claim_id="c1",
                            claim_text="=1+1",
                            source_ids=["src_formula"],
                            support_tier="tier_1_governing",
                            support_status="grounded",
                        )
                    ],
                    qa=QAResult(status="passed", notes=["grounded"]),
                    qa_grade="full",
                )
        ],
    )

    written = OutputWriter(tmp_path).write(artifacts)
    audit = json.loads(Path(written.output_paths["audit_json"]).read_text(encoding="utf-8"))
    payload = json.loads(Path(written.output_paths["json"]).read_text(encoding="utf-8"))

    assert _audit_json_value(audit, "Final Answer") == "=1+1"
    assert _audit_json_value(audit, "Evidence Summary") == "@SUM(A1:A2)"
    assert payload["answers"][0]["final_answer"] == "=1+1"
    assert payload["answers"][0]["evidence_summary"] == "@SUM(A1:A2)"


def test_output_writer_rejects_existing_run_without_overwriting(tmp_path):
    artifacts = RunArtifacts(
        run_id="run_existing",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 0},
        stats={"answered": 0, "empty": 0, "weak": 0, "failed": 0},
        answers=[],
    )
    writer = OutputWriter(tmp_path)
    written = writer.write(artifacts)
    json_path = Path(written.output_paths["json"])
    original = json_path.read_bytes()

    with pytest.raises(OutputRunExistsError, match="output run already exists"):
        writer.write(written)

    assert json_path.read_bytes() == original


def test_output_writer_temporal_retry_reuses_complete_existing_run(tmp_path):
    artifacts = RunArtifacts(
        run_id="run_retry",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 0},
        stats={"answered": 0, "empty": 0, "weak": 0, "failed": 0},
        answers=[],
    )
    writer = OutputWriter(tmp_path)
    first = writer.write(artifacts)
    json_path = Path(first.output_paths["json"])
    original = json_path.read_bytes()

    retried = writer.write(artifacts, retry_existing=True)

    assert retried.output_paths == first.output_paths
    assert json_path.read_bytes() == original
    assert not list(json_path.parent.glob("*.tmp*"))


@pytest.mark.parametrize("unsafe_value", ["../escape", r"..\escape", "bad/name"])
def test_output_writer_rejects_unsafe_artifact_identifiers(tmp_path, unsafe_value):
    artifacts = RunArtifacts(
        run_id="run_safe",
        company={"company_id": unsafe_value, "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 0},
        stats={"answered": 0, "empty": 0, "weak": 0, "failed": 0},
        answers=[],
    )

    with pytest.raises(ValueError):
        OutputWriter(tmp_path).write(artifacts)


def test_append_excel_row_uses_fallback_when_assignment_still_fails(tmp_path):
    class Unconvertible:
        pass

    writer = OutputWriter(tmp_path)

    wb = Workbook()
    ws = wb.active
    writer._append_excel_row(ws, [Unconvertible()])

    assert ws["A1"].value == EXCEL_FALLBACK_TEXT


def _combined_artifacts(run_id, company_name="C"):
    return RunArtifacts(
        run_id=run_id,
        company={"company_id": "c", "company_name": company_name, "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 1},
        stats={"answered": 1, "empty": 0, "weak": 0, "failed": 0},
        answers=[
            AnswerRecord(
                qid="Q001",
                source_id="EBX-Q-001",
                area="일반",
                category="ESG",
                question="question",
                answer_status="high_confidence",
                rag_pillar="governance",
                rag_retrieval_confidence=0.94,
                rag_coverage_status="complete",
                rag_answerable=True,
                rag_covered_facets=["accountable_body", "role"],
                final_answer="=unsafe",
                qa_grade="full",
                coverage_reason="complete_answer",
                original_evidence="accepted original evidence",
                evidence_summary="prompt evidence",
                sources=[{
                    "source_name": "report.pdf",
                    "source_path": "ESG/report.pdf",
                    "canonical_source_id": "src_report",
                    "source_tier": "tier_2_operational",
                    "document_status": "approved",
                }],
                qa=QAResult(status="passed", notes=[]),
                claim_support=[
                    ClaimSupport(
                        claim_id="c1",
                        claim_text="=unsafe",
                        source_ids=["src_report"],
                        support_tier="tier_2_operational",
                        support_status="grounded",
                    )
                ],
                skill_name="ESG writer",
                skill_version="1",
                skill_selection_reason="matched topic",
                raw_rag_result={
                    "items": [{"raw_evidence_ko": "raw rag evidence not used by writer"}],
                },
            )
        ],
        quantitative_results=[
            QuantitativeResult(
                metric_id=f"QUANT-{index:04d}",
                index=index,
                metric_name=f"Metric {index}",
                value=index if index == 1 else None,
                unit="unit",
                source="report.pdf" if index == 1 else "",
                status="filled" if index == 1 else "missing",
                confidence=0.99 if index == 1 else 0.0,
                metadata={"index": index},
            )
            for index in range(1, 252)
        ],
        quantitative_stats={"total": 251, "filled": 1, "missing": 250},
    )


def test_output_writer_creates_two_sheet_combined_workbook_and_numbered_names(tmp_path):
    now = lambda: datetime(2026, 7, 31, 9, 0, tzinfo=ZoneInfo("Asia/Bangkok"))
    writer = OutputWriter(tmp_path, now_provider=now)

    first = writer.write(_combined_artifacts("run_1", "Công ty: A/B"))
    second = writer.write(_combined_artifacts("run_2", "Công ty: A/B"))

    first_path = first.output_paths["combined_excel"]
    second_path = second.output_paths["combined_excel"]
    assert first_path.endswith("[langgraph][Công ty_ A_B]report-2026.07.31_1.xlsx")
    assert second_path.endswith("[langgraph][Công ty_ A_B]report-2026.07.31_2.xlsx")

    workbook = load_workbook(first_path, data_only=False)
    assert workbook.sheetnames == [
        "Qualitative",
        "Qualitative Table Metrics",
        "Quantitative",
    ]
    qualitative = workbook["Qualitative"]
    quantitative = workbook["Quantitative"]
    assert [cell.value for cell in qualitative[1]] == COMBINED_QUALITATIVE_COLUMNS
    assert [cell.value for cell in quantitative[1]] == QUANTITATIVE_COLUMNS
    assert qualitative["A2"].value == "EBX-Q-001"
    assert qualitative["B2"].value == "Answer: PUBLISHED\nEvidence: SUFFICIENT"
    assert qualitative["C2"].value == "일반 / 거버넌스 (Governance) / question"
    assert qualitative["D2"].value == "accepted original evidence"
    assert qualitative["E2"].value == "'=unsafe"
    assert quantitative.max_row == 252
    assert quantitative["A252"].value == "QUANT-0251"
    assert quantitative["D2"].value == 1
    assert quantitative.freeze_panes == "A2"
    first_run_dir = Path(first.output_paths["json"]).parent
    assert len(list(first_run_dir.iterdir())) == 4

    payload = json.loads(
        Path(first.output_paths["json"]).read_text(encoding="utf-8")
    )
    assert payload["quantitative_stats"] == {"total": 251, "filled": 1, "missing": 250}
    assert payload["output_paths"]["combined_excel"] == first_path


def test_coverage_summary_groups_failure_reasons():
    artifacts = RunArtifacts(
        run_id="run_coverage",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 4},
        stats={"answered": 1, "empty": 1, "weak": 1, "failed": 1},
        provenance={"mode": "strict", "verified": True, "source_digest": "abc123"},
        answers=[
            AnswerRecord(
                qid="Q001",
                result_bucket="weak",
                final_answer="",
                qa=QAResult(status="empty", notes=["empty evidence"]),
                retrieval_attempts=[{"top_k": 12, "retry_reason": "empty evidence", "eligible_item_count": 0}],
            ),
            AnswerRecord(
                qid="Q031",
                result_bucket="failed",
                final_answer="",
                qa=QAResult(
                    status="failed",
                    notes=["missing required facet: metric_result", "source usage overstated"],
                ),
                quality_flags=["draft_evidence"],
            ),
            AnswerRecord(
                qid="Q036",
                result_bucket="answered",
                final_answer="Partial cautious answer. The metric value was not disclosed.",
                qa=QAResult(status="passed", notes=["missing facet: target", "missing data disclosed"]),
                quality_flags=["partial_answer", "missing_facet:target", "draft_based_answer"],
            ),
            AnswerRecord(
                qid="Q051",
                result_bucket="failed",
                final_answer="",
                qa=QAResult(status="failed", notes=["unsupported certification or initiative claim"]),
            ),
        ],
    )

    summary = build_coverage_summary(artifacts)

    assert summary["total_qids"] == 4
    assert summary["bucket_stats"] == {"answered": 1, "empty": 1, "weak": 1, "failed": 1}
    assert summary["final_answer_stats"] == {"non_empty": 1, "empty": 3}
    assert summary["qa_stats"] == {"passed": 1, "failed": 2, "empty": 1}
    assert summary["quality_grade_stats"] == {
        "full": 0,
        "partial": 0,
        "cautious": 1,
        "failed": 3,
    }
    assert summary["quality_grade_qids"]["cautious"] == ["Q036"]
    assert summary["provenance"] == {
        "mode": "strict",
        "verified": True,
        "source_digest": "abc123",
    }
    assert summary["coverage_reason_qids"]["empty_evidence"] == ["Q001"]
    assert summary["coverage_reason_qids"]["draft_evidence"] == ["Q036"]
    assert summary["coverage_reason_qids"]["unsupported_claim"] == ["Q051"]
    assert summary["coverage_issue_qids"]["draft_evidence"] == ["Q031", "Q036"]
    assert sum(
        count
        for reasons in summary["coverage_matrix"].values()
        for count in reasons.values()
    ) == 4
    assert summary["empty_final_answer_qids"] == ["Q001", "Q031", "Q051"]
    assert summary["groups"]["empty_evidence"] == ["Q001"]
    assert summary["groups"]["metrics_missing"] == ["Q031"]
    assert summary["groups"]["draft_evidence"] == ["Q031", "Q036"]
    assert summary["groups"]["draft_based_answers"] == ["Q036"]
    assert summary["groups"]["partial_answers"] == ["Q036"]
    assert summary["groups"]["missing_required_facets"] == ["Q031"]
    assert summary["groups"]["missing_expected_facets"] == ["Q036"]
    assert summary["groups"]["source_overstated"] == ["Q031"]
    assert summary["groups"]["unsupported_claim"] == ["Q051"]
    assert summary["retrieval"]["retried_qids"] == ["Q001"]
    assert summary["retrieval"]["retry_unresolved_qids"] == ["Q001"]


def test_coverage_summary_reports_final_answer_counts_separately_from_buckets():
    answers = []
    for index in range(1, 96):
        final_answer = "answer" if index <= 30 else ""
        qa_status = "passed" if index <= 30 else "empty" if index <= 51 else "failed"
        bucket = "answered" if final_answer else "weak" if qa_status == "empty" else "failed"
        answers.append(
            AnswerRecord(
                qid=f"Q{index:03d}",
                result_bucket=bucket,
                final_answer=final_answer,
                qa=QAResult(status=qa_status, notes=[]),
            )
        )
    artifacts = RunArtifacts(
        run_id="run_coverage_counts",
        company={"company_id": "c", "company_name": "C", "year": 2025},
        template_selection={"template_version": "template_v1", "question_count": 95},
        stats={"answered": 30, "empty": 0, "weak": 21, "failed": 44},
        answers=answers,
    )

    summary = build_coverage_summary(artifacts)

    assert summary["stats"] == {"answered": 30, "empty": 0, "weak": 21, "failed": 44}
    assert summary["bucket_stats"] == summary["stats"]
    assert summary["final_answer_stats"] == {"non_empty": 30, "empty": 65}
    assert summary["qa_stats"] == {"passed": 30, "failed": 44, "empty": 21}
    assert summary["quality_grade_stats"] == {
        "full": 0,
        "partial": 0,
        "cautious": 30,
        "failed": 65,
    }
    assert sum(summary["quality_grade_stats"].values()) == 95
    assert len(summary["empty_final_answer_qids"]) == 65


def test_combined_filename_counter_is_unique_for_concurrent_runs(tmp_path):
    now = lambda: datetime(2026, 7, 31, 9, 0, tzinfo=ZoneInfo("Asia/Bangkok"))

    def write(run_id):
        return OutputWriter(tmp_path, now_provider=now).write(
            _combined_artifacts(run_id, "Concurrent Company")
        )

    with ThreadPoolExecutor(max_workers=2) as pool:
        results = list(pool.map(write, ["run_a", "run_b"]))

    filenames = {result.output_paths["combined_excel"].rsplit("\\", 1)[-1] for result in results}
    assert filenames == {
        "[langgraph][Concurrent Company]report-2026.07.31_1.xlsx",
        "[langgraph][Concurrent Company]report-2026.07.31_2.xlsx",
    }
