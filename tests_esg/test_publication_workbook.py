from __future__ import annotations

import json

from openpyxl import load_workbook

from esgagents.output_writer import AUDIT_COLUMNS, OutputWriter
from esgagents.schemas import AnswerRecord, QAResult, RunArtifacts


def test_review_answer_is_identical_in_json_audit_and_customer_workbook(tmp_path):
    candidate = "A supported answer that still needs human review."
    artifacts = RunArtifacts(
        run_id="publication_workbook",
        company={"company_id": "company", "company_name": "Company", "year": 2026},
        template_selection={"template_version": "v1", "question_count": 1},
        stats={"answered": 1, "empty": 0, "weak": 0, "failed": 0},
        answers=[
            AnswerRecord(
                qid="Q075",
                source_id="EBX-Q-075",
                answer_status="medium_confidence",
                final_answer=candidate,
                qa=QAResult(status="passed", notes=["missing expected metric dimension: activity_count"]),
                qa_grade="partial",
                coverage_reason="missing_metric_or_period",
                coverage_issues=["missing_metric_or_period"],
                publication_status="review_required",
                publication_reason="missing_metric_or_period",
                publication_issues=["qa_grade:partial"],
                rag_metric_evidence=[
                    {
                        "table_block": "Committee activity",
                        "raw_evidence_ko": "Committee meetings | 2025=4",
                        "facts": [{"metric": "Committee meetings", "period": "2025", "value": "4"}],
                    }
                ],
            )
        ],
    )

    written = OutputWriter(tmp_path).write(artifacts)
    payload = json.loads(
        open(written.output_paths["json"], encoding="utf-8").read()
    )
    audit = json.loads(open(written.output_paths["audit_json"], encoding="utf-8").read())
    customer = load_workbook(written.output_paths["combined_excel"])

    customer_answer = payload["answers"][0]["final_answer"]
    assert customer_answer == candidate
    assert "required metric or reporting period" not in customer_answer
    assert payload["answers"][0]["publication_status"] == "review_required"
    assert "missing_metric_or_period" in payload["answers"][0]["publication_issues"]
    assert audit["columns"] == AUDIT_COLUMNS
    assert audit["rows"][0]["Final Answer"] == customer_answer
    assert customer.sheetnames == ["Qualitative", "Qualitative Table Metrics"]
    assert customer["Qualitative"]["F2"].value == customer_answer
    assert customer["Qualitative"]["B2"].value == "Answer: REVIEW\nEvidence: METRIC_REVIEW"


def test_coverage_summary_contains_all_publication_buckets(tmp_path):
    answers = [
        AnswerRecord(
            qid="Q001",
            answer_status="high_confidence",
            final_answer="Published.",
            qa=QAResult(status="passed"),
            qa_grade="full",
            coverage_reason="complete_answer",
            publication_status="published",
            publication_reason="complete_grounded_answer",
        ),
        AnswerRecord(
            qid="Q002",
            answer_status="medium_confidence",
            final_answer="The company maintains a supported policy statement for customer review.",
            qa=QAResult(status="passed"),
            qa_grade="partial",
            coverage_reason="partial_answer",
            coverage_issues=["partial_answer"],
            publication_status="review_required",
            publication_reason="partial_answer",
            publication_issues=["qa_grade:partial"],
        ),
        AnswerRecord(
            qid="Q003",
            answer_status="insufficient",
            final_answer="",
            qa=QAResult(status="failed"),
            qa_grade="failed",
            coverage_reason="qa_failed",
            publication_status="blocked",
            publication_reason="unaccepted_answer_status",
            publication_issues=[
                "empty_final_answer",
                "unaccepted_answer_status:insufficient",
            ],
            rag_failure_code="NO_EVIDENCE",
        ),
    ]
    artifacts = RunArtifacts(
        run_id="publication_coverage",
        company={"company_id": "company", "company_name": "Company", "year": 2026},
        template_selection={"template_version": "v1", "question_count": 3},
        stats={"answered": 2, "empty": 0, "weak": 0, "failed": 1},
        answers=answers,
    )

    written = OutputWriter(tmp_path).write(artifacts)
    coverage = json.loads(
        open(written.output_paths["coverage_summary"], encoding="utf-8").read()
    )

    assert coverage["publication_status_stats"] == {
        "published": 1,
        "review_required": 1,
        "blocked": 1,
    }
    assert coverage["publication_status_qids"] == {
        "published": ["Q001"],
        "review_required": ["Q002"],
        "blocked": ["Q003"],
    }
    assert coverage["publication_issue_stats"] == {
        "empty_final_answer": 1,
        "partial_answer": 1,
        "qa_grade:partial": 1,
        "unaccepted_answer_status:insufficient": 1,
    }
    assert coverage["rag"]["upstream_insufficient"] == {
        "count": 1,
        "qids": ["Q003"],
        "failure_code_qids": {"NO_EVIDENCE": ["Q003"]},
    }
    assert coverage["customer_answer_count"] == 2
    assert coverage["customer_answer_qids"] == ["Q001", "Q002"]
    assert coverage["review_exported_count"] == 1
    assert coverage["review_exported_qids"] == ["Q002"]
    assert coverage["json_xlsx_answer_parity"] is True
    assert coverage["json_xlsx_answer_parity_mismatch_qids"] == []


def test_blocked_candidate_is_rejected_and_json_matches_customer_workbook(tmp_path):
    candidate = "In 2025, human-rights grievances totaled 63 cases."
    artifacts = RunArtifacts(
        run_id="publication_parity",
        company={"company_id": "company", "company_name": "Company", "year": 2026},
        template_selection={"template_version": "v1", "question_count": 2},
        stats={"answered": 2, "empty": 0, "weak": 0, "failed": 0},
        answers=[
            AnswerRecord(
                qid="Q011",
                source_id="EBX-Q-011",
                answer_status="high_confidence",
                final_answer=candidate,
                qa=QAResult(status="passed"),
                qa_grade="full",
                rag_metric_status="not_found",
                metric_audit={"metric_status": "not_found", "accepted_facts": []},
            ),
            AnswerRecord(
                qid="Q021",
                source_id="EBX-Q-021",
                answer_status="high_confidence",
                final_answer="The Board reviews and approves the EHS policy annually.",
                qa=QAResult(status="passed"),
                qa_grade="partial",
                coverage_reason="missing_expected_facets",
                coverage_issues=["missing_expected_facets"],
            ),
        ],
    )

    written = OutputWriter(tmp_path).write(artifacts)
    payload = json.loads(open(written.output_paths["json"], encoding="utf-8").read())
    customer = load_workbook(written.output_paths["combined_excel"])
    audit = json.loads(open(written.output_paths["audit_json"], encoding="utf-8").read())

    json_by_source = {answer["source_id"]: answer for answer in payload["answers"]}
    sheet = customer["Qualitative"]
    xlsx_by_source = {
        sheet.cell(row, 1).value: sheet.cell(row, 6).value or ""
        for row in range(2, sheet.max_row + 1)
    }
    assert all(
        json_by_source[source_id]["final_answer"] == xlsx_answer
        for source_id, xlsx_answer in xlsx_by_source.items()
    )
    assert json_by_source["EBX-Q-011"]["final_answer"] == ""
    assert json_by_source["EBX-Q-011"]["last_rejected_answer"] == candidate
    assert json_by_source["EBX-Q-021"]["final_answer"]
    assert audit["columns"] == AUDIT_COLUMNS
    assert audit["rows"][0]["Final Answer"] == candidate
    assert audit["rows"][0]["Last Rejected Answer"] == candidate
